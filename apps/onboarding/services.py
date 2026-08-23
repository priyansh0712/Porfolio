"""
Bulk Onboarding Service Layer — Dynamic Sample Template Generator, Parsers, Validation Engine & Atomic Committers.
"""
import csv
import io
import re
import logging
import datetime
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.db import transaction
from django.utils import timezone

from apps.accounts.models import User
from apps.faculty.models import Faculty, FacultyCustomField, FacultyFormFieldConfig
from apps.faculty.services import FacultyCodeService
from apps.academics.models import (
    AcademicYear, Standard, Division, Subject,
    ClassTeacherAllocation, SubjectTeacherAllocation, ClassCurriculum
)
from apps.students.models import Student, StudentCustomField, StudentFormFieldConfig
from apps.students.services import StudentService

logger = logging.getLogger(__name__)


class SampleTemplateService:
    """
    Generates pre-formatted downloadable sample templates (.xlsx & .csv)
    dynamically scoped to the school tenant's form field configuration & custom fields.
    """

    @classmethod
    def get_template_headers_and_data(cls, step, school=None):
        if step == 1:
            return cls._get_step_1_headers_and_data(school)
        elif step == 2:
            return cls._get_step_2_headers_and_data()
        elif step == 3:
            return cls._get_step_3_headers_and_data()
        elif step == 4:
            return cls._get_step_4_headers_and_data(school)
        return [], []

    @classmethod
    def _get_step_1_headers_and_data(cls, school=None):
        config = FacultyFormFieldConfig.get_for_school(school) if school else None

        # Core mandatory headers
        headers = ['First Name', 'Last Name', 'Email']
        sample_1 = ['John', 'Smith', 'john.smith@school.edu']
        sample_2 = ['Sarah', 'Connor', 'sarah.c@school.edu']

        # Dynamic Standard Fields based on FacultyFormFieldConfig
        if not config or config.show_employee_code:
            headers.append('Employee Code')
            sample_1.append('FAC-001')
            sample_2.append('FAC-002')

        if not config or config.show_department:
            headers.append('Department')
            sample_1.append('Science')
            sample_2.append('Mathematics')

        if not config or config.show_designation:
            headers.append('Designation')
            sample_1.append('Senior Physics Teacher')
            sample_2.append('Head of Maths')

        if not config or config.show_phone_number:
            headers.append('Phone Number')
            sample_1.append('9876543210')
            sample_2.append('9876543211')

        # Dynamic Custom Fields
        if school:
            custom_fields = list(FacultyCustomField.objects.filter(school=school, is_active=True).order_by('order_index', 'created_at'))
            for cf in custom_fields:
                headers.append(cf.label)
                if cf.field_type == FacultyCustomField.FieldType.NUMBER:
                    sample_1.append('100')
                    sample_2.append('101')
                elif cf.field_type == FacultyCustomField.FieldType.DATE:
                    sample_1.append('2026-01-01')
                    sample_2.append('2026-01-01')
                elif cf.field_type == FacultyCustomField.FieldType.SELECT:
                    opts = [o.strip() for o in cf.options.split(',') if o.strip()]
                    val = opts[0] if opts else 'Option1'
                    sample_1.append(val)
                    sample_2.append(val)
                else:
                    sample_1.append(f"Sample {cf.label}")
                    sample_2.append(f"Sample {cf.label}")

        return headers, [sample_1, sample_2]

    @classmethod
    def _get_step_2_headers_and_data(cls):
        headers = ['Standard Name', 'Standard Code', 'Division Name', 'Class Teacher Employee Code']
        sample_rows = [
            ['Grade 10', 'STD-10', 'A', 'FAC-001'],
            ['Grade 10', 'STD-10', 'B', 'FAC-002'],
        ]
        return headers, sample_rows

    @classmethod
    def _get_step_3_headers_and_data(cls):
        headers = ['Standard Name', 'Division Name', 'Subject Name', 'Subject Code', 'Subject Teacher Employee Code']
        sample_rows = [
            ['Grade 10', 'A', 'Physics', 'PHY-10', 'FAC-001'],
            ['Grade 10', 'A', 'Algebra', 'ALG-10', 'FAC-002'],
        ]
        return headers, sample_rows

    @classmethod
    def _get_step_4_headers_and_data(cls, school=None):
        config = StudentFormFieldConfig.get_for_school(school) if school else None

        # Core mandatory headers
        headers = ['GR Number', 'First Name', 'Last Name', 'Standard Name', 'Division Name']
        sample_1 = ['GR-1001', 'Alex', 'Taylor', 'Grade 10', 'A']
        sample_2 = ['GR-1002', 'Emily', 'Davis', 'Grade 10', 'A']

        # Dynamic Standard Fields based on StudentFormFieldConfig
        if not config or config.show_roll_number:
            headers.append('Roll Number')
            sample_1.append('1')
            sample_2.append('2')

        if not config or config.show_gender:
            headers.append('Gender')
            sample_1.append('Male')
            sample_2.append('Female')

        if not config or config.show_dob:
            headers.append('Date of Birth')
            sample_1.append('2010-05-15')
            sample_2.append('2010-08-20')

        if not config or config.show_blood_group:
            headers.append('Blood Group')
            sample_1.append('O+')
            sample_2.append('B+')

        if not config or config.show_guardian_details:
            headers.append('Guardian Name')
            sample_1.append('Robert Taylor')
            sample_2.append('Sarah Davis')

            headers.append('Parent Phone')
            sample_1.append('9876543210')
            sample_2.append('9876543211')

            headers.append('Parent Email')
            sample_1.append('parent1@gmail.com')
            sample_2.append('parent2@gmail.com')

        if not config or config.show_emergency_contact:
            headers.append('Emergency Contact')
            sample_1.append('9876543219')
            sample_2.append('9876543218')

        if not config or config.show_address:
            headers.append('Address')
            sample_1.append('123 MG Road, Ahmedabad')
            sample_2.append('456 Ring Road, Surat')

        # Dynamic Custom Fields
        if school:
            custom_fields = list(StudentCustomField.objects.filter(school=school, is_active=True).order_by('order_index', 'created_at'))
            for cf in custom_fields:
                headers.append(cf.label)
                if cf.field_type == StudentCustomField.FieldType.NUMBER:
                    sample_1.append('100')
                    sample_2.append('101')
                elif cf.field_type == StudentCustomField.FieldType.DATE:
                    sample_1.append('2026-01-01')
                    sample_2.append('2026-01-01')
                elif cf.field_type == StudentCustomField.FieldType.SELECT:
                    opts = [o.strip() for o in cf.options.split(',') if o.strip()]
                    val = opts[0] if opts else 'Option1'
                    sample_1.append(val)
                    sample_2.append(val)
                else:
                    sample_1.append(f"Sample {cf.label}")
                    sample_2.append(f"Sample {cf.label}")

        return headers, [sample_1, sample_2]

    @classmethod
    def generate_csv(cls, step, school=None):
        headers, sample_rows = cls.get_template_headers_and_data(step, school)

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in sample_rows:
            writer.writerow(row)
        return output.getvalue().encode('utf-8')

    @classmethod
    def generate_xlsx(cls, step, school=None):
        headers, sample_rows = cls.get_template_headers_and_data(step, school)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Step {step} Sample Import"

        # Styled Header Row (Apple Action Blue)
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 22

        # Data Rows
        for row in sample_rows:
            ws.append(row)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class BulkImportParser:
    """Parses .xlsx and .csv files into normalized row dictionaries."""

    @classmethod
    def _format_cell_value(cls, val):
        if val is None:
            return ''
        if isinstance(val, (datetime.datetime, datetime.date)):
            return val.strftime('%Y-%m-%d')
        if isinstance(val, float):
            if val.is_integer():
                return str(int(val))
            return str(val).strip()
        if isinstance(val, int):
            return str(val)
        return str(val).strip()

    @classmethod
    def parse(cls, file_obj, filename):
        filename_lower = filename.lower()
        if filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls'):
            return cls._parse_xlsx(file_obj)
        else:
            return cls._parse_csv(file_obj)

    @classmethod
    def _parse_xlsx(cls, file_obj):
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return []

        raw_headers = [str(h).strip() for h in rows[0] if h is not None]
        data_rows = []
        for r_idx, row in enumerate(rows[1:], start=2):
            if not any(cell is not None and str(cell).strip() != '' for cell in row):
                continue
            row_dict = {}
            for h_idx, h in enumerate(raw_headers):
                val = row[h_idx] if h_idx < len(row) else None
                row_dict[h] = cls._format_cell_value(val)
            data_rows.append((r_idx, row_dict))
        return data_rows

    @classmethod
    def _parse_csv(cls, file_obj):
        content = file_obj.read()
        try:
            text = content.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = content.decode('latin-1')

        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows or len(rows) < 2:
            return []

        headers = [h.strip() for h in rows[0] if h.strip()]
        data_rows = []
        for r_idx, row in enumerate(rows[1:], start=2):
            if not any(cell.strip() != '' for cell in row):
                continue
            row_dict = {}
            for h_idx, h in enumerate(headers):
                val = row[h_idx] if h_idx < len(row) else ''
                row_dict[h] = str(val).strip()
            data_rows.append((r_idx, row_dict))
        return data_rows


class BulkValidationService:
    """Multi-tenant scoped row-by-row validation engine."""

    @staticmethod
    def _validate_date_str(date_str):
        if not date_str:
            return None, True
        formats = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']
        for fmt in formats:
            try:
                d = datetime.datetime.strptime(date_str, fmt).date()
                return d, True
            except ValueError:
                pass
        return None, False

    @classmethod
    def _validate_custom_fields(cls, row, custom_fields, errors):
        for cf in custom_fields:
            raw_val = row.get(cf.label)
            val = str(raw_val).strip() if raw_val is not None else ''
            if cf.is_required and not val:
                errors.append(f"Custom Field '{cf.label}' is required")
                continue
            if not val:
                continue

            if cf.field_type == FacultyCustomField.FieldType.NUMBER:
                try:
                    float(val)
                except ValueError:
                    errors.append(f"Custom Field '{cf.label}' must be a valid number")
            elif cf.field_type == FacultyCustomField.FieldType.DATE:
                _, is_valid = cls._validate_date_str(val)
                if not is_valid:
                    errors.append(f"Custom Field '{cf.label}' must be a valid date (YYYY-MM-DD)")
            elif cf.field_type == FacultyCustomField.FieldType.SELECT:
                opts = [o.strip().lower() for o in cf.options.split(',') if o.strip()]
                if opts and val.lower() not in opts:
                    errors.append(f"Custom Field '{cf.label}' value '{val}' is invalid. Allowed: {cf.options}")

    @classmethod
    def validate(cls, school, step, raw_rows):
        if step == 1:
            return cls._validate_step_1_faculty(school, raw_rows)
        elif step == 2:
            return cls._validate_step_2_classes(school, raw_rows)
        elif step == 3:
            return cls._validate_step_3_subjects(school, raw_rows)
        elif step == 4:
            return cls._validate_step_4_students(school, raw_rows)
        return []

    @classmethod
    def _validate_step_1_faculty(cls, school, raw_rows):
        config = FacultyFormFieldConfig.get_for_school(school)
        existing_faculty_emails = set(Faculty.objects.filter(school=school).values_list('email', flat=True))
        existing_user_emails = set(User.objects.values_list('email', flat=True))
        existing_codes = set(Faculty.objects.filter(school=school).values_list('employee_code', flat=True))
        custom_fields = list(FacultyCustomField.objects.filter(school=school, is_active=True))

        seen_emails = set()
        seen_codes = set()
        results = []

        for r_idx, row in raw_rows:
            errors = []
            first_name = row.get('First Name', '').strip()
            last_name = row.get('Last Name', '').strip()
            email = row.get('Email', '').strip().lower()
            code = row.get('Employee Code', '').strip().upper()
            department = row.get('Department', '').strip()
            designation = row.get('Designation', '').strip()
            phone_number = row.get('Phone Number', '').strip()

            if not first_name:
                errors.append("First Name is required")
            if not last_name:
                errors.append("Last Name is required")
            if not email or '@' not in email:
                errors.append("Valid Email is required")
            elif email in existing_faculty_emails or email in existing_user_emails:
                errors.append(f"Email '{email}' is already registered in system")
            elif email in seen_emails:
                errors.append(f"Duplicate email '{email}' in import file")

            # Employee Code requirement
            if config.require_employee_code and not code:
                errors.append("Employee Code is required by school policy")
            elif code:
                if code in existing_codes:
                    errors.append(f"Employee Code '{code}' already exists")
                elif code in seen_codes:
                    errors.append(f"Duplicate Employee Code '{code}' in import file")

            # Department requirement
            if config.require_department and not department:
                errors.append("Department is required by school policy")

            # Designation requirement
            if config.require_designation and not designation:
                errors.append("Designation is required by school policy")

            # Phone number requirement
            if config.require_phone_number and not phone_number:
                errors.append("Phone Number is required by school policy")

            # Custom Field Validation
            cls._validate_custom_fields(row, custom_fields, errors)

            if email:
                seen_emails.add(email)
            if code:
                seen_codes.add(code)

            results.append({
                'row_index': r_idx,
                'data': row,
                'status': 'ERROR' if errors else 'VALID',
                'errors': errors,
            })
        return results

    @classmethod
    def _validate_step_2_classes(cls, school, raw_rows):
        curr_ay = AcademicYear.objects.filter(school=school, is_current=True).first()
        faculty_map = {f.employee_code.upper(): f for f in Faculty.objects.filter(school=school, is_active=True)}
        assigned_teachers = set(
            ClassTeacherAllocation.objects.filter(school=school, academic_year=curr_ay)
            .values_list('faculty__employee_code', flat=True)
        ) if curr_ay else set()

        seen_classes = set()
        seen_file_teachers = set()
        results = []

        for r_idx, row in raw_rows:
            errors = []
            std_name = row.get('Standard Name', '').strip()
            div_name = row.get('Division Name', '').strip()
            teacher_code = row.get('Class Teacher Employee Code', '').strip().upper()

            if not std_name:
                errors.append("Standard Name is required")
            if not div_name:
                errors.append("Division Name is required")

            class_key = (std_name.upper(), div_name.upper())
            if class_key in seen_classes:
                errors.append(f"Duplicate Class '{std_name} {div_name}' in import file")
            seen_classes.add(class_key)

            if teacher_code:
                if teacher_code not in faculty_map:
                    errors.append(f"Teacher Code '{teacher_code}' not found in school faculty roster")
                elif teacher_code in assigned_teachers:
                    errors.append(f"Teacher '{teacher_code}' is already assigned as Class Teacher in database")
                elif teacher_code in seen_file_teachers:
                    errors.append(f"Teacher '{teacher_code}' is assigned to multiple classes in this import file")
                else:
                    seen_file_teachers.add(teacher_code)

            results.append({
                'row_index': r_idx,
                'data': row,
                'status': 'ERROR' if errors else 'VALID',
                'errors': errors,
            })
        return results

    @classmethod
    def _validate_step_3_subjects(cls, school, raw_rows):
        curr_ay = AcademicYear.objects.filter(school=school, is_current=True).first()
        divisions = {
            (d.standard.name.upper(), d.name.upper()): d
            for d in Division.objects.filter(school=school).select_related('standard')
        }
        faculty_map = {f.employee_code.upper(): f for f in Faculty.objects.filter(school=school, is_active=True)}

        seen_allocations = set()
        results = []

        for r_idx, row in raw_rows:
            errors = []
            std_name = row.get('Standard Name', '').strip().upper()
            div_name = row.get('Division Name', '').strip().upper()
            sub_name = row.get('Subject Name', '').strip()
            teacher_code = row.get('Subject Teacher Employee Code', '').strip().upper()

            if not curr_ay:
                errors.append("No active Academic Year found for your school. Please activate an Academic Year first.")

            if not std_name or not div_name:
                errors.append("Standard Name and Division Name are required")
            elif (std_name, div_name) not in divisions:
                errors.append(f"Class '{std_name} {div_name}' does not exist. Please run Step 2 first.")

            if not sub_name:
                errors.append("Subject Name is required")

            if not teacher_code:
                errors.append("Subject Teacher Employee Code is required")
            elif teacher_code not in faculty_map:
                errors.append(f"Subject Teacher Code '{teacher_code}' not found in faculty roster")

            alloc_key = (std_name, div_name, sub_name.upper())
            if alloc_key in seen_allocations:
                errors.append(f"Duplicate Subject '{sub_name}' for Class '{std_name} {div_name}' in import file")
            seen_allocations.add(alloc_key)

            results.append({
                'row_index': r_idx,
                'data': row,
                'status': 'ERROR' if errors else 'VALID',
                'errors': errors,
            })
        return results

    @classmethod
    def _validate_step_4_students(cls, school, raw_rows):
        config = StudentFormFieldConfig.get_for_school(school)
        curr_ay = AcademicYear.objects.filter(school=school, is_current=True).first()
        existing_grs = set(Student.objects.filter(school=school).values_list('gr_number', flat=True))
        divisions = {
            (d.standard.name.upper(), d.name.upper()): d
            for d in Division.objects.filter(school=school).select_related('standard')
        }
        custom_fields = list(StudentCustomField.objects.filter(school=school, is_active=True))

        # Query existing roll numbers in DB per division for current academic year
        existing_rolls_by_div = {}
        if curr_ay:
            existing_records = Student.objects.filter(
                school=school,
                academic_year=curr_ay,
                is_active=True,
                roll_number__isnull=False
            ).values_list('division_id', 'roll_number')
            for div_id, r_no in existing_records:
                existing_rolls_by_div.setdefault(div_id, set()).add(r_no)

        seen_grs = set()
        seen_rolls = set()
        results = []

        for r_idx, row in raw_rows:
            errors = []
            gr_number = str(row.get('GR Number', '') or '').strip().upper()
            first_name = str(row.get('First Name', '') or '').strip()
            last_name = str(row.get('Last Name', '') or '').strip()
            std_name = str(row.get('Standard Name', '') or '').strip().upper()
            div_name = str(row.get('Division Name', '') or '').strip().upper()
            roll_str = str(row.get('Roll Number', '') or '').strip()
            dob_str = str(row.get('Date of Birth', '') or '').strip()
            gender_str = str(row.get('Gender', '') or '').strip()
            blood_group = str(row.get('Blood Group', '') or '').strip()
            guardian_name = str(row.get('Guardian Name', '') or '').strip()
            parent_phone = str(row.get('Parent Phone', '') or '').strip()
            emergency_contact = str(row.get('Emergency Contact', '') or '').strip()
            address = str(row.get('Address', '') or '').strip()

            if not curr_ay:
                errors.append("No active Academic Year found. Please activate an Academic Year before importing students.")

            if not gr_number:
                errors.append("GR Number is required")
            elif gr_number in existing_grs:
                errors.append(f"GR Number '{gr_number}' already exists in school")
            elif gr_number in seen_grs:
                errors.append(f"Duplicate GR Number '{gr_number}' in file")

            if not first_name or not last_name:
                errors.append("First Name and Last Name are required")

            div_obj = None
            if not std_name or not div_name:
                errors.append("Standard Name and Division Name are required")
            elif (std_name, div_name) not in divisions:
                errors.append(f"Class '{std_name} {div_name}' does not exist. Please run Step 2 first.")
            else:
                div_obj = divisions[(std_name, div_name)]

            # Roll number validation
            if config.require_roll_number and not roll_str:
                errors.append("Roll Number is required by school policy")
            elif roll_str:
                try:
                    roll_num = int(float(roll_str))
                    if roll_num <= 0:
                        errors.append("Roll Number must be a positive integer")
                    elif div_obj:
                        if roll_num in existing_rolls_by_div.get(div_obj.id, set()):
                            errors.append(f"Roll Number '{roll_num}' already exists in Class {std_name} {div_name}")
                        roll_key = (div_obj.id, roll_num)
                        if roll_key in seen_rolls:
                            errors.append(f"Duplicate Roll Number '{roll_num}' in Class {std_name} {div_name} in import file")
                        seen_rolls.add(roll_key)
                except ValueError:
                    errors.append(f"Invalid Roll Number '{roll_str}'")

            # Gender validation
            if config.require_gender and not gender_str:
                errors.append("Gender is required by school policy")
            elif gender_str:
                normalized_gender = gender_str.upper()
                if normalized_gender in ('M', 'BOY', 'MALE'):
                    pass
                elif normalized_gender in ('F', 'GIRL', 'FEMALE'):
                    pass
                elif normalized_gender in ('O', 'OTHER'):
                    pass
                else:
                    errors.append(f"Invalid Gender '{gender_str}'. Allowed: Male, Female, Other")

            # Date of Birth validation
            if config.require_dob and not dob_str:
                errors.append("Date of Birth is required by school policy")
            elif dob_str:
                _, is_valid = cls._validate_date_str(dob_str)
                if not is_valid:
                    errors.append(f"Invalid Date of Birth '{dob_str}'. Expected format: YYYY-MM-DD")

            # Blood Group validation
            if config.require_blood_group and not blood_group:
                errors.append("Blood Group is required by school policy")

            # Guardian details validation
            if config.require_guardian_details and not guardian_name:
                errors.append("Guardian Name is required by school policy")
            if config.require_guardian_details and not parent_phone:
                errors.append("Parent Phone is required by school policy")

            # Emergency contact
            if config.require_emergency_contact and not emergency_contact:
                errors.append("Emergency Contact is required by school policy")

            # Address
            if config.require_address and not address:
                errors.append("Address is required by school policy")

            # Custom Field Validation
            cls._validate_custom_fields(row, custom_fields, errors)

            if gr_number:
                seen_grs.add(gr_number)

            results.append({
                'row_index': r_idx,
                'data': row,
                'status': 'ERROR' if errors else 'VALID',
                'errors': errors,
            })
        return results


class BulkCommitService:
    """Executes atomic database transactions for validated rows."""

    @classmethod
    @transaction.atomic
    def commit_step_1_faculty(cls, school, valid_rows, default_password='Admin@123'):
        custom_fields = list(FacultyCustomField.objects.filter(school=school, is_active=True))

        count = 0
        for item in valid_rows:
            row = item['data']
            first_name = str(row['First Name']).strip()
            last_name = str(row['Last Name']).strip()
            email = str(row['Email']).strip().lower()
            code = str(row.get('Employee Code', '') or '').strip().upper()
            if not code:
                code = FacultyCodeService.generate_next_code(school)

            department = str(row.get('Department', '') or '').strip()
            designation = str(row.get('Designation', '') or '').strip()
            phone_number = str(row.get('Phone Number', '') or '').strip()

            # Extract Custom Field Values
            custom_data = {}
            for cf in custom_fields:
                val = row.get(cf.label)
                if val is not None and str(val).strip() != '':
                    custom_data[cf.field_name] = str(val).strip()

            faculty = Faculty.objects.create(
                school=school,
                first_name=first_name,
                last_name=last_name,
                email=email,
                employee_code=code,
                department=department,
                designation=designation,
                phone_number=phone_number,
                custom_fields=custom_data,
                is_active=True,
            )

            # Auto-create User account
            user = User.objects.create_user(
                username=email,
                email=email,
                password=default_password,
                first_name=first_name,
                last_name=last_name,
                role=User.Role.FACULTY,
                school=school,
            )
            faculty.user = user
            faculty.save(update_fields=['user'])
            count += 1
        return count

    @classmethod
    @transaction.atomic
    def commit_step_2_classes(cls, school, valid_rows):
        start_date = timezone.now().date()
        end_date = start_date + datetime.timedelta(days=365)
        curr_ay, _ = AcademicYear.objects.get_or_create(
            school=school,
            is_current=True,
            defaults={'name': f'{start_date.year}-{start_date.year + 1}', 'start_date': start_date, 'end_date': end_date}
        )
        faculty_map = {f.employee_code.upper(): f for f in Faculty.objects.filter(school=school, is_active=True)}
        count = 0

        for item in valid_rows:
            row = item['data']
            std_name = row['Standard Name'].strip()
            std_code = row.get('Standard Code') or std_name.upper().replace(' ', '-')
            div_name = row['Division Name'].strip()
            teacher_code = str(row.get('Class Teacher Employee Code', '') or '').strip().upper()

            standard, _ = Standard.objects.get_or_create(
                school=school,
                name=std_name
            )

            division, _ = Division.objects.get_or_create(
                school=school,
                standard=standard,
                name=div_name
            )

            if teacher_code and teacher_code in faculty_map:
                faculty = faculty_map[teacher_code]
                ClassTeacherAllocation.objects.update_or_create(
                    school=school,
                    academic_year=curr_ay,
                    division=division,
                    defaults={'faculty': faculty}
                )
            count += 1
        return count

    @classmethod
    @transaction.atomic
    def commit_step_3_subjects(cls, school, valid_rows):
        curr_ay = AcademicYear.objects.filter(school=school, is_current=True).first()
        faculty_map = {f.employee_code.upper(): f for f in Faculty.objects.filter(school=school, is_active=True)}
        divisions = {
            (d.standard.name.upper(), d.name.upper()): d
            for d in Division.objects.filter(school=school).select_related('standard')
        }

        count = 0
        for item in valid_rows:
            row = item['data']
            std_name = row['Standard Name'].strip().upper()
            div_name = row['Division Name'].strip().upper()
            sub_name = row['Subject Name'].strip()
            sub_code = str(row.get('Subject Code', '') or '').strip().upper() or sub_name.upper()[:6]
            teacher_code = str(row['Subject Teacher Employee Code']).strip().upper()

            div_obj = divisions.get((std_name, div_name))
            if not div_obj:
                continue

            subject, _ = Subject.objects.get_or_create(
                school=school,
                code=sub_code,
                defaults={'name': sub_name}
            )

            # Maintain ClassCurriculum entry
            ClassCurriculum.objects.get_or_create(
                school=school,
                academic_year=curr_ay,
                standard=div_obj.standard,
                subject=subject,
                defaults={'is_active': True}
            )

            faculty = faculty_map[teacher_code]
            SubjectTeacherAllocation.objects.update_or_create(
                school=school,
                academic_year=curr_ay,
                division=div_obj,
                subject=subject,
                faculty=faculty,
            )
            count += 1
        return count

    @classmethod
    @transaction.atomic
    def commit_step_4_students(cls, school, valid_rows, default_password='Admin@123'):
        curr_ay = AcademicYear.objects.filter(school=school, is_current=True).first()
        divisions = {
            (d.standard.name.upper(), d.name.upper()): d
            for d in Division.objects.filter(school=school).select_related('standard')
        }
        custom_fields = list(StudentCustomField.objects.filter(school=school, is_active=True))

        count = 0
        for item in valid_rows:
            row = item['data']
            gr_number = str(row.get('GR Number', '') or '').strip().upper()
            first_name = str(row.get('First Name', '') or '').strip()
            last_name = str(row.get('Last Name', '') or '').strip()
            full_name = f"{first_name} {last_name}".strip()
            std_name = str(row.get('Standard Name', '') or '').strip().upper()
            div_name = str(row.get('Division Name', '') or '').strip().upper()

            div_obj = divisions.get((std_name, div_name))
            if not div_obj:
                continue

            # Parse DOB safely
            dob = None
            dob_raw = row.get('Date of Birth')
            if dob_raw:
                dob, _ = BulkValidationService._validate_date_str(str(dob_raw).strip())

            # Parse Roll Number safely
            roll_num = None
            roll_raw = row.get('Roll Number')
            if roll_raw is not None and str(roll_raw).strip() != '':
                try:
                    roll_num = int(float(str(roll_raw).strip()))
                except ValueError:
                    roll_num = None

            # Normalize Gender
            raw_gender = str(row.get('Gender', 'MALE')).strip().upper()
            if raw_gender in ('F', 'FEMALE', 'GIRL'):
                gender = 'FEMALE'
            elif raw_gender in ('O', 'OTHER'):
                gender = 'OTHER'
            else:
                gender = 'MALE'

            blood_group = str(row.get('Blood Group', '') or '').strip()
            guardian_name = str(row.get('Guardian Name', '') or '').strip()
            guardian_phone = str(row.get('Parent Phone', '') or '').strip()
            emergency_contact = str(row.get('Emergency Contact', '') or '').strip()
            address = str(row.get('Address', '') or '').strip()

            # Extract Custom Field Values
            custom_data = {}
            for cf in custom_fields:
                val = row.get(cf.label)
                if val is not None and str(val).strip() != '':
                    custom_data[cf.field_name] = str(val).strip()

            StudentService.create_student(
                school=school,
                academic_year=curr_ay,
                standard=div_obj.standard,
                division=div_obj,
                gr_number=gr_number,
                full_name=full_name,
                roll_number=roll_num,
                dob=dob,
                gender=gender,
                blood_group=blood_group,
                guardian_name=guardian_name,
                guardian_phone=guardian_phone,
                emergency_contact=emergency_contact,
                address=address,
                custom_fields=custom_data,
            )
            count += 1
        return count
