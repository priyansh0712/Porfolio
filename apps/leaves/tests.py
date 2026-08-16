from django.test import TestCase
from django.core.exceptions import ValidationError
from django.db.utils import IntegrityError
from django.core.files.uploadedfile import SimpleUploadedFile
from datetime import date

from apps.accounts.models import User
from apps.faculty.models import Faculty
from apps.tenants.models import School
from apps.leaves.models import LeaveAllocation, LeaveRequest, LeaveType
from apps.leaves.services import LeaveAllocationService
from apps.leaves.forms import LeaveRequestForm
from apps.notifications.models import InAppNotification
from apps.attendance.models import AttendanceLog
from apps.schedules.models import WorkingSchedule, HolidayException
from django.utils import timezone
from apps.tenants.context import set_current_tenant, _current_tenant


class LeavesModelTests(TestCase):

    @classmethod
    def setUpTestData(cls):
        # Create two school tenants
        cls.school_a = School.objects.create(
            name="Alpha School",
            subdomain="alpha",
            contact_email="admin@alpha.edu",
            is_active=True
        )
        cls.school_b = School.objects.create(
            name="Beta School",
            subdomain="beta",
            contact_email="admin@beta.edu",
            is_active=True
        )

        # Create user and faculty for School A
        cls.user_a = User.objects.create_user(
            username="faculty_a@alpha.edu",
            email="faculty_a@alpha.edu",
            password="FacultyPass1!",
            role=User.Role.FACULTY,
            school=cls.school_a
        )
        cls.faculty_a = Faculty.objects.create(
            school=cls.school_a,
            user=cls.user_a,
            first_name="John",
            last_name="Doe",
            email="faculty_a@alpha.edu",
            employee_code="ALPHA-001",
            department="Science"
        )

        # Create user and faculty for School B
        cls.user_b = User.objects.create_user(
            username="faculty_b@beta.edu",
            email="faculty_b@beta.edu",
            password="FacultyPass1!",
            role=User.Role.FACULTY,
            school=cls.school_b
        )
        cls.faculty_b = Faculty.objects.create(
            school=cls.school_b,
            user=cls.user_b,
            first_name="Jane",
            last_name="Smith",
            email="faculty_b@beta.edu",
            employee_code="BETA-001",
            department="Math"
        )

    def tearDown(self):
        # Reset tenant context to None to avoid leaking state
        set_current_tenant(None)

    def test_create_leave_allocation(self):
        """Verify leave allocation creation and string representation."""
        alloc = LeaveAllocation.objects.create(
            school=self.school_a,
            faculty=self.faculty_a,
            leave_type=LeaveType.CASUAL,
            allocated=12
        )
        self.assertEqual(alloc.allocated, 12)
        self.assertEqual(alloc.leave_type, LeaveType.CASUAL)
        self.assertEqual(str(alloc), "John Doe — Casual Leave: 12")

    def test_unique_allocation_constraint(self):
        """Verify that a duplicate leave allocation per faculty + leave type raises IntegrityError."""
        LeaveAllocation.objects.create(
            school=self.school_a,
            faculty=self.faculty_a,
            leave_type=LeaveType.CASUAL,
            allocated=12
        )
        with self.assertRaises(IntegrityError):
            LeaveAllocation.objects.create(
                school=self.school_a,
                faculty=self.faculty_a,
                leave_type=LeaveType.CASUAL,
                allocated=10
            )

    def test_create_leave_request(self):
        """Verify leave request creation, default status, and string representation."""
        req = LeaveRequest.objects.create(
            school=self.school_a,
            faculty=self.faculty_a,
            leave_type=LeaveType.SICK,
            from_date=date(2026, 9, 1),
            to_date=date(2026, 9, 3),
            reason="Medical checkup"
        )
        self.assertEqual(req.status, LeaveRequest.Status.PENDING)
        self.assertEqual(req.reason, "Medical checkup")
        self.assertEqual(str(req), "John Doe — Sick Leave (2026-09-01 to 2026-09-03)")

    def test_leave_request_date_validation(self):
        """Verify that from_date > to_date raises ValidationError."""
        req = LeaveRequest(
            school=self.school_a,
            faculty=self.faculty_a,
            leave_type=LeaveType.CASUAL,
            from_date=date(2026, 9, 5),
            to_date=date(2026, 9, 3),
            reason="Invalid dates"
        )
        with self.assertRaises(ValidationError):
            req.full_clean()

    def test_tenant_scoping_isolation(self):
        """Verify that TenantManager automatically filters leave records by active tenant."""
        # Create allocations for School A and School B
        LeaveAllocation.objects.create(
            school=self.school_a,
            faculty=self.faculty_a,
            leave_type=LeaveType.CASUAL,
            allocated=12
        )
        LeaveAllocation.objects.create(
            school=self.school_b,
            faculty=self.faculty_b,
            leave_type=LeaveType.CASUAL,
            allocated=10
        )

        # Set tenant context to School A
        set_current_tenant(self.school_a)
        allocations_a = LeaveAllocation.objects.all()
        self.assertEqual(allocations_a.count(), 1)
        self.assertEqual(allocations_a.first().faculty, self.faculty_a)

        # Set tenant context to School B
        set_current_tenant(self.school_b)
        allocations_b = LeaveAllocation.objects.all()
        self.assertEqual(allocations_b.count(), 1)
        self.assertEqual(allocations_b.first().faculty, self.faculty_b)


class LeaveAllocationIntegrationTests(TestCase):

    @classmethod
    def setUpTestData(cls):
        # Create school tenants
        cls.school_a = School.objects.create(
            name="Alpha School",
            subdomain="alpha",
            contact_email="admin@alpha.edu",
            is_active=True
        )
        cls.school_b = School.objects.create(
            name="Beta School",
            subdomain="beta",
            contact_email="admin@beta.edu",
            is_active=True
        )

        # Create School Admin for School A
        cls.admin_a = User.objects.create_user(
            username="admin_a@alpha.edu",
            email="admin_a@alpha.edu",
            password="AdminPass1!",
            role=User.Role.SCHOOL_ADMIN,
            school=cls.school_a
        )

        # Create Faculty for School A
        cls.faculty_a = Faculty.objects.create(
            school=cls.school_a,
            first_name="John",
            last_name="Doe",
            email="faculty_a@alpha.edu",
            employee_code="ALPHA-001",
            department="Science"
        )

        # Create Faculty for School B
        cls.faculty_b = Faculty.objects.create(
            school=cls.school_b,
            first_name="Jane",
            last_name="Smith",
            email="faculty_b@beta.edu",
            employee_code="BETA-001",
            department="Math"
        )

    def tearDown(self):
        set_current_tenant(None)

    def _create_excel_file(self, rows):
        """Helper to construct an Excel spreadsheet in memory."""
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leave Allocations"
        for r in rows:
            ws.append(r)
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    def test_generate_excel_template(self):
        """Verify dynamic template contains active faculty member code."""
        import io
        import openpyxl
        data = LeaveAllocationService.generate_excel_template(self.school_a)
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        
        # Verify columns headers
        headers = [c.value for c in ws[1]]
        self.assertIn("Faculty ID", headers)
        self.assertIn("Casual Leave", headers)
        
        # Verify prepopulated row
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(len(rows), 2)  # Header + 1 Faculty
        self.assertEqual(rows[1][0], "ALPHA-001")

    def test_download_template_view_permissions(self):
        """Verify School Admin can download template, unauthenticated is redirected."""
        # Unauthenticated
        response = self.client.get("/leaves/allocation/template/")
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Authenticated Admin
        self.client.force_login(self.admin_a)
        # Mock TenantMiddleware resolution
        session = self.client.session
        session['tenant_subdomain'] = 'alpha'
        session.save()
        
        response = self.client.get("/leaves/allocation/template/", HTTP_HOST="alpha.localhost")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_upload_valid_excel(self):
        """Verify importing valid excel updates database allocations successfully."""
        self.client.force_login(self.admin_a)
        
        file_stream = self._create_excel_file([
            ["Faculty ID", "Faculty Name", "Casual Leave", "Sick Leave", "Paid Leave"],
            ["ALPHA-001", "John Doe", "15", "8", "20"]
        ])
        uploaded_file = SimpleUploadedFile("allocs.xlsx", file_stream.read(), content_type="application/vnd.ms-excel")

        response = self.client.post(
            "/leaves/allocation/upload/",
            {"file": uploaded_file},
            HTTP_HOST="alpha.localhost"
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("success_message", response.context)
        
        # Verify DB values
        allocs = {
            a.leave_type: a.allocated
            for a in LeaveAllocation.objects.filter(school=self.school_a, faculty=self.faculty_a)
        }
        self.assertEqual(allocs[LeaveType.CASUAL], 15)
        self.assertEqual(allocs[LeaveType.SICK], 8)
        self.assertEqual(allocs[LeaveType.PAID], 20)

    def test_upload_invalid_excel_rolls_back_entire_transaction(self):
        """Verify that any row error triggers an atomic rollback (all or nothing)."""
        self.client.force_login(self.admin_a)
        
        file_stream = self._create_excel_file([
            ["Faculty ID", "Faculty Name", "Casual Leave", "Sick Leave", "Paid Leave"],
            ["ALPHA-001", "John Doe", "15", "-2", "20"]  # Row contains negative number
        ])
        uploaded_file = SimpleUploadedFile("allocs_invalid.xlsx", file_stream.read(), content_type="application/vnd.ms-excel")

        response = self.client.post(
            "/leaves/allocation/upload/",
            {"file": uploaded_file},
            HTTP_HOST="alpha.localhost"
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("row_errors", response.context)
        self.assertIn("Row 2: Sick Leave - Value must be a positive/non-negative integer.", response.context["row_errors"])
        
        # Verify DB remained unchanged (no partial saves)
        self.assertEqual(LeaveAllocation.objects.filter(school=self.school_a, faculty=self.faculty_a).count(), 0)

    def test_upload_cross_tenant_validation_block(self):
        """Verify that uploading an Excel containing a foreign tenant's Faculty ID is rejected."""
        self.client.force_login(self.admin_a)
        
        file_stream = self._create_excel_file([
            ["Faculty ID", "Faculty Name", "Casual Leave", "Sick Leave", "Paid Leave"],
            ["BETA-001", "Jane Smith", "15", "8", "20"]  # Faculty belongs to School B
        ])
        uploaded_file = SimpleUploadedFile("allocs_cross.xlsx", file_stream.read(), content_type="application/vnd.ms-excel")

        response = self.client.post(
            "/leaves/allocation/upload/",
            {"file": uploaded_file},
            HTTP_HOST="alpha.localhost"
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("row_errors", response.context)
        self.assertIn("Row 2: Faculty ID 'BETA-001' does not exist in your school.", response.context["row_errors"])
        
        # Verify no allocations were written for either tenant
        self.assertEqual(LeaveAllocation.objects.filter(faculty=self.faculty_b).count(), 0)


class FacultyDashboardAndLeaveTests(TestCase):

    @classmethod
    def setUpTestData(cls):
        # Create school tenants
        cls.school_a = School.objects.create(
            name="Alpha School",
            subdomain="alpha",
            contact_email="admin@alpha.edu",
            is_active=True
        )
        cls.school_b = School.objects.create(
            name="Beta School",
            subdomain="beta",
            contact_email="admin@beta.edu",
            is_active=True
        )

        # Create School Admin for School A
        cls.admin_a = User.objects.create_user(
            username="admin_a@alpha.edu",
            email="admin_a@alpha.edu",
            password="AdminPass1!",
            role=User.Role.SCHOOL_ADMIN,
            school=cls.school_a
        )

        # Create Faculty with password for School A
        from apps.faculty.services import FacultyService
        cls.faculty_a = FacultyService.create_faculty(
            school=cls.school_a,
            data={
                "first_name": "John",
                "last_name": "Doe",
                "email": "john@alpha.edu",
                "department": "Science",
                "password": "FacultyPass1!"
            }
        )

    def tearDown(self):
        set_current_tenant(None)

    def test_faculty_login_succeeds(self):
        """Verify that a Faculty member can authenticate using their password."""
        login_data = {
            "username": "john@alpha.edu",
            "password": "FacultyPass1!"
        }
        # Attempt login on the correct subdomain
        login_success = self.client.login(username="john@alpha.edu", password="FacultyPass1!", HTTP_HOST="alpha.localhost")
        self.assertTrue(login_success)

    def test_dashboard_unified_routing(self):
        """Verify GET /dashboard/ routes to Admin/Faculty dashboard templates by user role."""
        session = self.client.session
        session['tenant_subdomain'] = 'alpha'
        session.save()

        # 1. School Admin
        self.client.force_login(self.admin_a)
        response = self.client.get("/dashboard/", HTTP_HOST="alpha.localhost")
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "reports/dashboard.html")

        # 2. Faculty User
        self.client.force_login(self.faculty_a.user)
        response = self.client.get("/dashboard/", HTTP_HOST="alpha.localhost")
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "leaves/faculty_dashboard.html")

    def test_apply_leave_success(self):
        """Verify valid leave request submission creates PENDING database entry."""
        self.client.force_login(self.faculty_a.user)

        # Create allocation so balance check passes
        LeaveAllocation.objects.create(
            school=self.school_a,
            faculty=self.faculty_a,
            leave_type=LeaveType.CASUAL,
            allocated=10
        )

        response = self.client.post(
            "/leaves/apply/",
            {
                "leave_type": LeaveType.CASUAL,
                "from_date": "2026-09-01",
                "to_date": "2026-09-03",
                "reason": "Family gathering"
            },
            HTTP_HOST="alpha.localhost"
        )
        self.assertRedirects(response, "/dashboard/")

        # Verify DB entry
        req = LeaveRequest.objects.filter(faculty=self.faculty_a).first()
        self.assertIsNotNone(req)
        self.assertEqual(req.leave_type, LeaveType.CASUAL)
        self.assertEqual(req.status, LeaveRequest.Status.PENDING)
        self.assertEqual(req.from_date, date(2026, 9, 1))
        self.assertEqual(req.to_date, date(2026, 9, 3))

    def test_apply_leave_date_ordering_validation(self):
        """Verify error is raised when from_date is after to_date."""
        self.client.force_login(self.faculty_a.user)

        response = self.client.post(
            "/leaves/apply/",
            {
                "leave_type": LeaveType.CASUAL,
                "from_date": "2026-09-05",
                "to_date": "2026-09-03",
                "reason": "Invalid range"
            },
            HTTP_HOST="alpha.localhost"
        )
        self.assertEqual(response.status_code, 200)
        form = response.context["form"]
        self.assertIn("from_date", form.errors)
        self.assertEqual(form.errors["from_date"][0], "From Date cannot be after To Date.")

    def test_apply_leave_overlapping_validation(self):
        """Verify overlapping approved/pending requests are blocked."""
        self.client.force_login(self.faculty_a.user)

        # Create allocation
        LeaveAllocation.objects.create(
            school=self.school_a,
            faculty=self.faculty_a,
            leave_type=LeaveType.CASUAL,
            allocated=10
        )

        # Create existing pending request
        LeaveRequest.objects.create(
            school=self.school_a,
            faculty=self.faculty_a,
            leave_type=LeaveType.CASUAL,
            from_date=date(2026, 9, 1),
            to_date=date(2026, 9, 3),
            reason="Existing leave",
            status=LeaveRequest.Status.PENDING
        )

        # Attempt to post overlapping request
        response = self.client.post(
            "/leaves/apply/",
            {
                "leave_type": LeaveType.CASUAL,
                "from_date": "2026-09-02",
                "to_date": "2026-09-04",
                "reason": "Conflicting leave"
            },
            HTTP_HOST="alpha.localhost"
        )
        self.assertEqual(response.status_code, 200)
        form = response.context["form"]
        self.assertIn("__all__", form.errors)
        self.assertEqual(form.errors["__all__"][0], "This leave request overlaps with another pending or approved request.")

    def test_apply_leave_balance_insufficient_validation(self):
        """Verify that requesting leaves exceeding allocated balance is rejected."""
        self.client.force_login(self.faculty_a.user)

        # Set allocation = 5 days
        LeaveAllocation.objects.create(
            school=self.school_a,
            faculty=self.faculty_a,
            leave_type=LeaveType.CASUAL,
            allocated=5
        )

        # Attempt to apply for 6 working days (Sep 1 to Sep 8)
        response = self.client.post(
            "/leaves/apply/",
            {
                "leave_type": LeaveType.CASUAL,
                "from_date": "2026-09-01",
                "to_date": "2026-09-08",
                "reason": "Exceeds balance"
            },
            HTTP_HOST="alpha.localhost"
        )
        self.assertEqual(response.status_code, 200)
        form = response.context["form"]
        self.assertIn("__all__", form.errors)
        self.assertIn("Insufficient leave balance for Casual Leave.", form.errors["__all__"][0])


class AdminLeaveReviewAndActionTests(TestCase):

    @classmethod
    def setUpTestData(cls):
        # Create school tenants
        cls.school_a = School.objects.create(
            name="Alpha School",
            subdomain="alpha",
            contact_email="admin@alpha.edu",
            is_active=True
        )
        cls.school_b = School.objects.create(
            name="Beta School",
            subdomain="beta",
            contact_email="admin@beta.edu",
            is_active=True
        )

        # Create School Admin for School A
        cls.admin_a = User.objects.create_user(
            username="admin_a@alpha.edu",
            email="admin_a@alpha.edu",
            password="AdminPass1!",
            role=User.Role.SCHOOL_ADMIN,
            school=cls.school_a
        )

        # Create School Admin for School B
        cls.admin_b = User.objects.create_user(
            username="admin_b@beta.edu",
            email="admin_b@beta.edu",
            password="AdminPass1!",
            role=User.Role.SCHOOL_ADMIN,
            school=cls.school_b
        )

        # Create Faculty for School A
        from apps.faculty.services import FacultyService
        cls.faculty_a = FacultyService.create_faculty(
            school=cls.school_a,
            data={
                "first_name": "John",
                "last_name": "Doe",
                "email": "john@alpha.edu",
                "employee_code": "ALPHA-001",
                "department": "Science"
            }
        )

        # Create Faculty for School B
        cls.faculty_b = FacultyService.create_faculty(
            school=cls.school_b,
            data={
                "first_name": "Jane",
                "last_name": "Smith",
                "email": "jane@beta.edu",
                "employee_code": "BETA-001",
                "department": "Math"
            }
        )

        # Create Leave Requests
        cls.req_a = LeaveRequest.objects.create(
            school=cls.school_a,
            faculty=cls.faculty_a,
            leave_type=LeaveType.CASUAL,
            from_date=date(2026, 9, 1),
            to_date=date(2026, 9, 3),
            reason="Family trip",
            status=LeaveRequest.Status.PENDING
        )

        cls.req_b = LeaveRequest.objects.create(
            school=cls.school_b,
            faculty=cls.faculty_b,
            leave_type=LeaveType.SICK,
            from_date=date(2026, 9, 5),
            to_date=date(2026, 9, 6),
            reason="Sick leave",
            status=LeaveRequest.Status.PENDING
        )

    def tearDown(self):
        set_current_tenant(None)

    def test_admin_requests_access_permissions(self):
        """Verify only School Admin can access requests page, others blocked/redirected."""
        # Unauthenticated redirects to login
        response = self.client.get("/leaves/admin/requests/")
        self.assertEqual(response.status_code, 302)

        # Faculty blocked (403)
        self.client.force_login(self.faculty_a.user)
        response = self.client.get("/leaves/admin/requests/", HTTP_HOST="alpha.localhost")
        self.assertEqual(response.status_code, 403)

        # Admin allowed (200)
        self.client.force_login(self.admin_a)
        response = self.client.get("/leaves/admin/requests/", HTTP_HOST="alpha.localhost")
        self.assertEqual(response.status_code, 200)

    def test_admin_requests_filtering(self):
        """Verify list filters return correct subset of leave requests."""
        self.client.force_login(self.admin_a)

        # Base requests check (should only see school A requests)
        response = self.client.get("/leaves/admin/requests/", HTTP_HOST="alpha.localhost")
        self.assertEqual(len(response.context["leave_requests"]), 1)
        self.assertEqual(response.context["leave_requests"][0], self.req_a)

        # Filter by status = PENDING (1 match)
        response = self.client.get("/leaves/admin/requests/?status=PENDING", HTTP_HOST="alpha.localhost")
        self.assertEqual(len(response.context["leave_requests"]), 1)

        # Filter by status = APPROVED (0 matches)
        response = self.client.get("/leaves/admin/requests/?status=APPROVED", HTTP_HOST="alpha.localhost")
        self.assertEqual(len(response.context["leave_requests"]), 0)

    def test_approve_leave_request_success(self):
        """Verify approving pending request transitions status and raises notification."""
        self.client.force_login(self.admin_a)

        response = self.client.post(
            f"/leaves/requests/{self.req_a.pk}/approve/",
            HTTP_HOST="alpha.localhost"
        )
        self.assertRedirects(response, "/leaves/admin/requests/")

        # Verify DB updates
        self.req_a.refresh_from_db()
        self.assertEqual(self.req_a.status, LeaveRequest.Status.APPROVED)

        # Verify InAppNotification creation
        notif = InAppNotification.objects.filter(user=self.faculty_a.user).first()
        self.assertIsNotNone(notif)
        self.assertEqual(notif.title, "Leave Request Approved")
        self.assertIn("approved", notif.message)

    def test_approve_already_processed_fails(self):
        """Verify cannot approve a request that is not pending."""
        self.client.force_login(self.admin_a)
        self.req_a.status = LeaveRequest.Status.APPROVED
        self.req_a.save()

        response = self.client.post(
            f"/leaves/requests/{self.req_a.pk}/approve/",
            HTTP_HOST="alpha.localhost"
        )
        self.assertRedirects(response, "/leaves/admin/requests/")
        
        # Verify status did not change
        self.req_a.refresh_from_db()
        self.assertEqual(self.req_a.status, LeaveRequest.Status.APPROVED)

    def test_reject_leave_request_success(self):
        """Verify rejecting request with reason transitions status and raises notification."""
        self.client.force_login(self.admin_a)

        response = self.client.post(
            f"/leaves/requests/{self.req_a.pk}/reject/",
            {"rejection_reason": "Insufficient project backup coverage."},
            HTTP_HOST="alpha.localhost"
        )
        self.assertRedirects(response, "/leaves/admin/requests/")

        # Verify DB updates
        self.req_a.refresh_from_db()
        self.assertEqual(self.req_a.status, LeaveRequest.Status.REJECTED)
        self.assertEqual(self.req_a.rejection_reason, "Insufficient project backup coverage.")

        # Verify InAppNotification creation
        notif = InAppNotification.objects.filter(user=self.faculty_a.user).first()
        self.assertIsNotNone(notif)
        self.assertEqual(notif.title, "Leave Request Rejected")
        self.assertIn("Insufficient project backup coverage", notif.message)

    def test_reject_missing_reason_fails(self):
        """Verify that rejecting a request without reason is blocked."""
        self.client.force_login(self.admin_a)

        response = self.client.post(
            f"/leaves/requests/{self.req_a.pk}/reject/",
            {"rejection_reason": ""},  # Blank reason
            HTTP_HOST="alpha.localhost"
        )
        self.assertRedirects(response, "/leaves/admin/requests/")

        # Verify DB unchanged (remains PENDING)
        self.req_a.refresh_from_db()
        self.assertEqual(self.req_a.status, LeaveRequest.Status.PENDING)

    def test_cross_tenant_approve_reject_blocked(self):
        """Verify that School Admin cannot approve/reject requests from another school."""
        self.client.force_login(self.admin_a)  # School A Admin

        # Attempt to approve School B's request (req_b)
        response = self.client.post(
            f"/leaves/requests/{self.req_b.pk}/approve/",
            HTTP_HOST="alpha.localhost"
        )
        self.assertEqual(response.status_code, 404)

        # Attempt to reject School B's request (req_b)
        response = self.client.post(
            f"/leaves/requests/{self.req_b.pk}/reject/",
            {"rejection_reason": "No access"},
            HTTP_HOST="alpha.localhost"
        )
        self.assertEqual(response.status_code, 404)

        # Verify School B request remains PENDING
        self.req_b.refresh_from_db()
        self.assertEqual(self.req_b.status, LeaveRequest.Status.PENDING)


class LeaveAttendanceIntegrationTests(TestCase):

    @classmethod
    def setUpTestData(cls):
        cls.school_a = School.objects.create(
            name="Alpha School",
            subdomain="alpha",
            contact_email="admin@alpha.edu",
            is_active=True
        )

        from apps.faculty.services import FacultyService
        cls.faculty_a = FacultyService.create_faculty(
            school=cls.school_a,
            data={
                "first_name": "John",
                "last_name": "Doe",
                "email": "john@alpha.edu",
                "employee_code": "ALPHA-001",
                "department": "Science"
            }
        )

    def tearDown(self):
        set_current_tenant(None)

    def test_leave_approval_generates_and_reverses_attendance_logs(self):
        """Verify approving a request auto-writes LEAVE logs; cancelling deletes them."""
        # Create request
        req = LeaveRequest.objects.create(
            school=self.school_a,
            faculty=self.faculty_a,
            leave_type=LeaveType.CASUAL,
            from_date=date(2026, 9, 10),
            to_date=date(2026, 9, 12),
            reason="Vacation",
            status=LeaveRequest.Status.PENDING
        )

        # Assert no logs exist initially
        logs = AttendanceLog.objects.filter(faculty=self.faculty_a, date__range=(date(2026, 9, 10), date(2026, 9, 12)))
        self.assertEqual(logs.count(), 0)

        # Approve the request
        req.status = LeaveRequest.Status.APPROVED
        req.save()

        # Verify logs generated
        logs = AttendanceLog.objects.filter(faculty=self.faculty_a, date__range=(date(2026, 9, 10), date(2026, 9, 12)))
        self.assertEqual(logs.count(), 3)
        for log in logs:
            self.assertEqual(log.status, AttendanceLog.Status.LEAVE)
            self.assertIsNone(log.check_in_time)
            self.assertIsNone(log.check_out_time)

        # Cancel request
        req.status = LeaveRequest.Status.CANCELLED
        req.save()

        # Verify logs deleted
        logs = AttendanceLog.objects.filter(faculty=self.faculty_a, date__range=(date(2026, 9, 10), date(2026, 9, 12)))
        self.assertEqual(logs.count(), 0)

    def test_dashboard_service_reflects_leave_count(self):
        """Verify dashboard KPIs correctly count today's leaves and adjust absent calculations."""
        today = timezone.localdate()

        # Create leave log for today
        AttendanceLog.objects.create(
            school=self.school_a,
            faculty=self.faculty_a,
            date=today,
            status=AttendanceLog.Status.LEAVE,
            last_scan_at=timezone.now()
        )

        from apps.reports.services import DashboardService
        metrics = DashboardService.get_metrics(self.school_a)

        self.assertEqual(metrics["total_scans"], 0)


class FacultyAttendanceHistoryTests(TestCase):

    @classmethod
    def setUpTestData(cls):
        # Create school tenants
        cls.school_a = School.objects.create(
            name="Alpha School",
            subdomain="alpha",
            contact_email="admin@alpha.edu",
            is_active=True
        )
        cls.school_b = School.objects.create(
            name="Beta School",
            subdomain="beta",
            contact_email="admin@beta.edu",
            is_active=True
        )

        # Create School Admin for School A
        cls.admin_a = User.objects.create_user(
            username="admin_a@alpha.edu",
            password="AdminPass1!",
            role=User.Role.SCHOOL_ADMIN,
            school=cls.school_a
        )

        # Create Faculty for School A
        from apps.faculty.services import FacultyService
        cls.faculty_a = FacultyService.create_faculty(
            school=cls.school_a,
            data={
                "first_name": "John",
                "last_name": "Doe",
                "email": "john@alpha.edu",
                "employee_code": "ALPHA-001",
                "department": "Science",
                "password": "FacultyPass1!"
            }
        )

        # Create Faculty for School B
        cls.faculty_b = FacultyService.create_faculty(
            school=cls.school_b,
            data={
                "first_name": "Jane",
                "last_name": "Smith",
                "email": "jane@beta.edu",
                "employee_code": "BETA-001",
                "department": "Math",
                "password": "FacultyPass1!"
            }
        )

        # Create Attendance Logs for Faculty A
        AttendanceLog.objects.create(
            school=cls.school_a,
            faculty=cls.faculty_a,
            date=date(2026, 9, 1),
            status=AttendanceLog.Status.PRESENT,
            last_scan_at=timezone.now()
        )
        AttendanceLog.objects.create(
            school=cls.school_a,
            faculty=cls.faculty_a,
            date=date(2026, 9, 2),
            status=AttendanceLog.Status.LATE,
            last_scan_at=timezone.now()
        )
        AttendanceLog.objects.create(
            school=cls.school_a,
            faculty=cls.faculty_a,
            date=date(2026, 9, 3),
            status=AttendanceLog.Status.LEAVE,
            last_scan_at=timezone.now()
        )

    def tearDown(self):
        set_current_tenant(None)

    def test_my_attendance_access_permissions(self):
        """Verify only Faculty role can view their personal attendance page."""
        # Unauthenticated redirects to login
        response = self.client.get("/leaves/my-attendance/")
        self.assertEqual(response.status_code, 302)

        # School Admin blocked
        self.client.force_login(self.admin_a)
        response = self.client.get("/leaves/my-attendance/", HTTP_HOST="alpha.localhost")
        self.assertEqual(response.status_code, 403)

        # Faculty allowed
        self.client.force_login(self.faculty_a.user)
        response = self.client.get("/leaves/my-attendance/", HTTP_HOST="alpha.localhost")
        self.assertEqual(response.status_code, 200)

    def test_my_attendance_summaries_calculation(self):
        """Verify counts for Present, Late, Leave, and Absent are computed accurately."""
        self.client.force_login(self.faculty_a.user)

        # Request range spanning 5 days (Sep 1 to Sep 5)
        response = self.client.get(
            "/leaves/my-attendance/?from_date=2026-09-01&to_date=2026-09-05",
            HTTP_HOST="alpha.localhost"
        )
        self.assertEqual(response.status_code, 200)
        
        # Verify Context metrics
        self.assertEqual(response.context["present_count"], 1)
        self.assertEqual(response.context["late_count"], 1)
        self.assertEqual(response.context["leave_count"], 1)
        # 5 days in range minus 3 logs = 2 absent days
        self.assertEqual(response.context["absent_count"], 2)

    def test_cross_tenant_isolation_on_attendance_history(self):
        """Verify Faculty B cannot see logs or counts belonging to Faculty A."""
        self.client.force_login(self.faculty_b.user)

        response = self.client.get(
            "/leaves/my-attendance/?from_date=2026-09-01&to_date=2026-09-05",
            HTTP_HOST="beta.localhost"
        )
        self.assertEqual(response.status_code, 200)
        
        # Verify Faculty B sees 0 metrics for school A logs
        self.assertEqual(response.context["present_count"], 0)
        self.assertEqual(response.context["late_count"], 0)
        self.assertEqual(response.context["leave_count"], 0)
        self.assertEqual(response.context["absent_count"], 5)  # 5 days total, 0 logs
        self.assertEqual(len(response.context["page_obj"]), 0)


class LeaveHolidayScheduleIntegrationTests(TestCase):

    @classmethod
    def setUpTestData(cls):
        # Create school tenant
        cls.school_a = School.objects.create(
            name="Alpha School",
            subdomain="alpha",
            contact_email="admin@alpha.edu",
            is_active=True
        )

        # Create Faculty user
        from apps.faculty.services import FacultyService
        cls.faculty_a = FacultyService.create_faculty(
            school=cls.school_a,
            data={
                "first_name": "John",
                "last_name": "Doe",
                "email": "john@alpha.edu",
                "employee_code": "ALPHA-001",
                "department": "Science",
                "password": "FacultyPass1!"
            }
        )

        # Setup standard 5-day WorkingSchedule (Mon-Fri)
        for d_idx in range(7):
            WorkingSchedule.objects.create(
                school=cls.school_a,
                day_of_week=d_idx,
                is_working_day=(d_idx < 5),  # Saturday/Sunday are non-working
                start_time="08:00:00",
                end_time="16:00:00",
                grace_period_minutes=15
            )

        # Setup custom holiday exception on Sep 15, 2026 (Tuesday)
        HolidayException.objects.create(
            school=cls.school_a,
            date=date(2026, 9, 15),
            description="Independence Day Holiday",
            is_recurring_yearly=False
        )

        # Setup recurring holiday exception on Sep 25 (Friday)
        HolidayException.objects.create(
            school=cls.school_a,
            date=date(2026, 9, 25),
            description="Yearly Festivity",
            is_recurring_yearly=True
        )

        # Allocate 10 CASUAL leave days to John Doe
        cls.alloc = LeaveAllocation.objects.create(
            school=cls.school_a,
            faculty=cls.faculty_a,
            leave_type=LeaveType.CASUAL,
            allocated=10
        )

    def tearDown(self):
        set_current_tenant(None)

    def test_leave_request_used_days_excludes_weekends_and_holidays(self):
        """Verify dynamic LeaveRequest used_days ignores weekend and holiday dates."""
        # 1. Sep 14 (Mon) to Sep 18 (Fri) spans Tue Sep 15 (holiday)
        # Expected: 4 days (Mon, Wed, Thu, Fri)
        req = LeaveRequest(
            school=self.school_a,
            faculty=self.faculty_a,
            leave_type=LeaveType.CASUAL,
            from_date=date(2026, 9, 14),
            to_date=date(2026, 9, 18)
        )
        self.assertEqual(req.used_days, 4)

        # 2. Sep 11 (Fri) to Sep 15 (Tue) spans weekend (Sat/Sun) + Tue (holiday)
        # Expected: 2 days (Fri, Mon)
        req2 = LeaveRequest(
            school=self.school_a,
            faculty=self.faculty_a,
            leave_type=LeaveType.CASUAL,
            from_date=date(2026, 9, 11),
            to_date=date(2026, 9, 15)
        )
        self.assertEqual(req2.used_days, 2)

        # 3. Sep 24 (Thu) to Sep 26 (Sat) spans Fri Sep 25 (recurring holiday)
        # Expected: 1 day (Thu)
        req3 = LeaveRequest(
            school=self.school_a,
            faculty=self.faculty_a,
            leave_type=LeaveType.CASUAL,
            from_date=date(2026, 9, 24),
            to_date=date(2026, 9, 26)
        )
        self.assertEqual(req3.used_days, 1)

    def test_validation_blocks_weekend_only_request(self):
        """Verify submitting a leave request spanning only off-days fails validation."""
        form_data = {
            "leave_type": LeaveType.CASUAL,
            "from_date": "2026-09-12",  # Saturday
            "to_date": "2026-09-13",    # Sunday
            "reason": "Weekend rest"
        }
        form = LeaveRequestForm(
            data=form_data,
            faculty=self.faculty_a,
            school=self.school_a
        )
        self.assertFalse(form.is_valid())
        self.assertIn("__all__", form.errors)
        self.assertEqual(
            form.errors["__all__"][0],
            "The selected date range does not contain any scheduled working days."
        )

    def test_balance_check_uses_schedule_adjusted_days(self):
        """Verify validation correctly checks remaining balance using working days."""
        # Remaining balance is 10. Let's make approved requests totaling 8 used days.
        # Request spans Sep 1 (Tue) to Sep 10 (Thu): 10 calendar days.
        # Weekend days: Sep 5 (Sat), Sep 6 (Sun). Net working days: 8.
        # This request uses 8 days.
        approved_req = LeaveRequest.objects.create(
            school=self.school_a,
            faculty=self.faculty_a,
            leave_type=LeaveType.CASUAL,
            from_date=date(2026, 9, 1),
            to_date=date(2026, 9, 10),
            status=LeaveRequest.Status.APPROVED
        )
        self.assertEqual(approved_req.used_days, 8)

        # Remaining balance is now: 10 - 8 = 2 days.
        # Attempt to apply for Sep 14 (Mon) to Sep 18 (Fri) - 4 working days.
        # 4 > 2, so it should fail validation.
        form_data_invalid = {
            "leave_type": LeaveType.CASUAL,
            "from_date": "2026-09-14",
            "to_date": "2026-09-18",
            "reason": "Need 4 days"
        }
        form_invalid = LeaveRequestForm(
            data=form_data_invalid,
            faculty=self.faculty_a,
            school=self.school_a
        )
        self.assertFalse(form_invalid.is_valid())
        self.assertIn("__all__", form_invalid.errors)
        self.assertIn("Insufficient leave balance", form_invalid.errors["__all__"][0])

        # Apply for Sep 14 (Mon) to Sep 15 (Tue) - Tuesday is holiday, so 1 working day.
        # 1 <= 2, so it should succeed.
        form_data_valid = {
            "leave_type": LeaveType.CASUAL,
            "from_date": "2026-09-14",
            "to_date": "2026-09-15",
            "reason": "Need 1 day"
        }
        form_valid = LeaveRequestForm(
            data=form_data_valid,
            faculty=self.faculty_a,
            school=self.school_a
        )
        self.assertTrue(form_valid.is_valid())






