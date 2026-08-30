"""
Timetable Business Services — Conflict Detection Engine & Excel Bulk Import/Export.
"""
import io
import re
from datetime import time, datetime
from typing import Dict, List, Optional, Tuple, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db import transaction
from django.db.models import Q
from django.core.exceptions import ValidationError

from apps.accounts.models import User
from apps.academics.models import (
    AcademicYear, Standard, Division, Subject, ClassTimetable, ClassTeacherAllocation
)
from apps.faculty.models import Faculty


class TimetableService:
    """
    Business logic for Timetable validation, conflict checking, and permission resolution.
    """

    DAY_MAPPING = {
        'monday': 1, 'mon': 1, '1': 1,
        'tuesday': 2, 'tue': 2, '2': 2,
        'wednesday': 3, 'wed': 3, '3': 3,
        'thursday': 4, 'thu': 4, '4': 4,
        'friday': 5, 'fri': 5, '5': 5,
        'saturday': 6, 'sat': 6, '6': 6,
    }

    @classmethod
    def can_manage_division_timetable(cls, user: User, division: Division, academic_year: AcademicYear) -> bool:
        """
        Determines if a user has authorization to manage the timetable for a specific Division.
        - Super Admins and School Admins can manage all divisions in their school.
        - Faculty members can manage ONLY if they are the allocated Class Teacher for this division and year.
        """
        if not user or not user.is_authenticated:
            return False

        if user.role in [User.Role.SUPER_ADMIN, User.Role.SCHOOL_ADMIN]:
            return division.school_id == user.school_id

        if user.role == User.Role.FACULTY:
            if division.school_id != user.school_id:
                return False
            return ClassTeacherAllocation.objects.filter(
                school=user.school,
                academic_year=academic_year,
                division=division,
                faculty__user=user,
            ).exists()

        return False

    @classmethod
    def get_manageable_divisions(cls, user: User, school, academic_year: AcademicYear):
        """
        Returns the queryset of divisions that the user is authorized to manage timetables for.
        """
        if user.role in [User.Role.SUPER_ADMIN, User.Role.SCHOOL_ADMIN]:
            return Division.objects.filter(
                school=school,
                is_active=True,
            ).select_related('standard').order_by('standard__order_index', 'name')

        if user.role == User.Role.FACULTY:
            assigned_div_ids = ClassTeacherAllocation.objects.filter(
                school=school,
                academic_year=academic_year,
                faculty__user=user,
            ).values_list('division_id', flat=True)

            return Division.objects.filter(
                school=school,
                is_active=True,
                pk__in=assigned_div_ids,
            ).select_related('standard').order_by('standard__order_index', 'name')

        return Division.objects.none()

    @classmethod
    def parse_time_str(cls, val: Any) -> Optional[time]:
        """
        Safely converts strings or datetime/time objects into a python datetime.time.
        Supports '08:00', '8:00', '08:00:00', '8:00 AM', etc.
        """
        if not val:
            return None
        if isinstance(val, time):
            return val
        if isinstance(val, datetime):
            return val.time()

        val_str = str(val).strip()
        formats = ['%H:%M:%S', '%H:%M', '%I:%M %p', '%I:%M%p', '%I:%M:%S %p']
        for fmt in formats:
            try:
                dt = datetime.strptime(val_str, fmt)
                return dt.time()
            except ValueError:
                continue
        return None

    @classmethod
    def validate_slot_conflicts(
        cls,
        school,
        academic_year: AcademicYear,
        division: Division,
        day_of_week: int,
        period_number: int,
        faculty: Optional[Faculty] = None,
        start_time: Optional[time] = None,
        end_time: Optional[time] = None,
        exclude_slot_id: Optional[int] = None,
    ) -> List[str]:
        """
        Validates timetable slot for conflicts:
        1. Time validity (end_time > start_time)
        2. Faculty double-booking across different classes on the same day/period or overlapping time.
        """
        errors = []
        day_name = dict(ClassTimetable.DayOfWeek.choices).get(day_of_week, f"Day {day_of_week}")

        # 1. Time Sequence Check
        if start_time and end_time:
            if end_time <= start_time:
                errors.append(
                    f"Invalid period timing: End time ({end_time.strftime('%H:%M')}) must be after start time ({start_time.strftime('%H:%M')})."
                )

        # 2. Faculty Double-Booking Check
        if faculty:
            # Query all other timetable slots for this faculty on this day
            faculty_slots_qs = ClassTimetable.objects.filter(
                school=school,
                academic_year=academic_year,
                faculty=faculty,
                day_of_week=day_of_week,
            ).exclude(division=division).select_related('division__standard', 'subject')

            if exclude_slot_id:
                faculty_slots_qs = faculty_slots_qs.exclude(pk=exclude_slot_id)

            for other_slot in faculty_slots_qs:
                conflicting_class_name = f"{other_slot.division.standard.name} - {other_slot.division.name}"

                # Conflict A: Same Period
                if other_slot.period_number == period_number:
                    errors.append(
                        f"Faculty conflict: Teacher '{faculty.full_name}' is already assigned to {conflicting_class_name} "
                        f"on {day_name} Period {period_number} ({other_slot.subject.name})."
                    )
                    continue

                # Conflict B: Overlapping Times
                if start_time and end_time and other_slot.start_time and other_slot.end_time:
                    # Overlap if max(start1, start2) < min(end1, end2)
                    if max(start_time, other_slot.start_time) < min(end_time, other_slot.end_time):
                        errors.append(
                            f"Faculty time overlap: Teacher '{faculty.full_name}' is scheduled in {conflicting_class_name} "
                            f"from {other_slot.start_time.strftime('%H:%M')} to {other_slot.end_time.strftime('%H:%M')} on {day_name}."
                        )

        return errors


class TimetableExcelService:
    """
    Service for generating downloadable Timetable Excel template and processing bulk uploads.
    """

    @classmethod
    def generate_sample_template(cls, school, academic_year: AcademicYear) -> bytes:
        """
        Creates an Apple-styled Excel workbook containing:
        1. 'Timetable Upload' sheet with required headers and example rows.
        2. 'Reference Masters' sheet with valid Classes, Subjects, and Faculty for the school.
        """
        wb = openpyxl.Workbook()

        # Sheet 1: Main Upload Sheet
        ws = wb.active
        ws.title = "Timetable Upload"

        headers = ["Class", "Day", "Period", "Start Time", "End Time", "Subject", "Faculty"]
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0'),
        )

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
            ws.row_dimensions[1].height = 28

        # Get existing divisions, subjects, faculty for realistic samples
        divisions = Division.objects.filter(school=school, is_active=True).select_related('standard')
        subjects = Subject.objects.filter(school=school, is_active=True)
        faculties = Faculty.objects.filter(school=school, is_active=True)

        sample_class_1 = f"{divisions[0].standard.name}-{divisions[0].name}" if divisions.exists() else "10-A"
        sample_class_2 = f"{divisions[1].standard.name}-{divisions[1].name}" if len(divisions) > 1 else sample_class_1
        sample_sub_1 = subjects[0].name if subjects.exists() else "Mathematics"
        sample_sub_2 = subjects[1].name if len(subjects) > 1 else "Science"
        sample_fac_1 = faculties[0].full_name if faculties.exists() else "Teacher 1"
        sample_fac_2 = faculties[1].full_name if len(faculties) > 1 else "Teacher 2"

        sample_rows = [
            [sample_class_1, "Monday", 1, "08:00", "08:45", sample_sub_1, sample_fac_1],
            [sample_class_1, "Monday", 2, "08:45", "09:30", sample_sub_2, sample_fac_2],
            [sample_class_1, "Tuesday", 1, "08:00", "08:45", sample_sub_2, sample_fac_2],
            [sample_class_2, "Monday", 1, "08:00", "08:45", sample_sub_2, sample_fac_2],
        ]

        sample_font = Font(name="Arial", size=10)
        for row in sample_rows:
            ws.append(row)
            row_idx = ws.max_row
            ws.row_dimensions[row_idx].height = 22
            for col_num in range(1, len(row) + 1):
                c = ws.cell(row=row_idx, column=col_num)
                c.font = sample_font
                c.border = border
                c.alignment = center_align if col_num in [2, 3, 4, 5] else left_align

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        # Sheet 2: Reference Masters
        ws_ref = wb.create_sheet(title="Reference Masters")
        ref_header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")

        ref_headers = ["Valid Classes", "Valid Subjects", "Valid Teachers (Faculty)"]
        ws_ref.append(ref_headers)
        ws_ref.row_dimensions[1].height = 26
        for col_num in range(1, 4):
            c = ws_ref.cell(row=1, column=col_num)
            c.font = header_font
            c.fill = ref_header_fill
            c.alignment = center_align
            c.border = border

        div_names = [f"{d.standard.name}-{d.name}" for d in divisions]
        sub_names = [s.name for s in subjects]
        fac_names = [f.full_name for f in faculties]

        max_rows = max(len(div_names), len(sub_names), len(fac_names), 1)
        for i in range(max_rows):
            c_val = div_names[i] if i < len(div_names) else ""
            s_val = sub_names[i] if i < len(sub_names) else ""
            f_val = fac_names[i] if i < len(fac_names) else ""
            ws_ref.append([c_val, s_val, f_val])
            r_idx = ws_ref.max_row
            ws_ref.row_dimensions[r_idx].height = 20
            for col_idx in range(1, 4):
                cell = ws_ref.cell(row=r_idx, column=col_idx)
                cell.font = sample_font
                cell.border = border

        for col in ws_ref.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_ref.column_dimensions[col_letter].width = max(max_len + 4, 20)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @classmethod
    def import_timetable_excel(
        cls,
        school,
        academic_year: AcademicYear,
        file_obj,
        user: User,
    ) -> Dict[str, Any]:
        """
        Parses and validates an uploaded Timetable Excel file.
        Returns:
            {
                'total_processed': int,
                'successful': int,
                'failed': int,
                'errors': List[str],
            }
        """
        filename = getattr(file_obj, 'name', 'upload.xlsx').lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            return {
                'total_processed': 0,
                'successful': 0,
                'failed': 0,
                'errors': ["Invalid file type. Please upload a valid Microsoft Excel (.xlsx or .xls) file."],
            }

        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
        except Exception as e:
            return {
                'total_processed': 0,
                'successful': 0,
                'failed': 0,
                'errors': [f"Could not open Excel file: {str(e)}"],
            }

        # 1. Map Header Columns
        header_row = [str(cell.value or '').strip().lower() for cell in ws[1]]
        col_map = {}
        for idx, col_name in enumerate(header_row):
            clean_name = re.sub(r'[^a-z0-9]', '', col_name)
            if 'class' in clean_name or 'division' in clean_name:
                col_map['class'] = idx
            elif 'day' in clean_name:
                col_map['day'] = idx
            elif 'period' in clean_name:
                col_map['period'] = idx
            elif 'start' in clean_name:
                col_map['start_time'] = idx
            elif 'end' in clean_name:
                col_map['end_time'] = idx
            elif 'subject' in clean_name:
                col_map['subject'] = idx
            elif 'faculty' in clean_name or 'teacher' in clean_name:
                col_map['faculty'] = idx

        required_cols = ['class', 'day', 'period', 'subject']
        missing = [c.capitalize() for c in required_cols if c not in col_map]
        if missing:
            return {
                'total_processed': 0,
                'successful': 0,
                'failed': 0,
                'errors': [f"Missing required columns in Excel sheet: {', '.join(missing)}. Please use the downloadable template."],
            }

        # 2. Build In-Memory Lookups (Strictly Scoped to Current School Tenant)
        divisions = Division.objects.filter(school=school, is_active=True).select_related('standard')
        division_lookup: Dict[str, Division] = {}
        for d in divisions:
            # Map standard name + division name, e.g. '10-a', 'standard 10-a', 'class 10-a', '10a'
            std_name = d.standard.name.lower().strip()
            div_name = d.name.lower().strip()
            division_lookup[f"{std_name}-{div_name}"] = d
            division_lookup[f"{std_name} {div_name}"] = d
            division_lookup[f"{d.standard.order_index}-{div_name}"] = d
            division_lookup[f"{d.standard.order_index}{div_name}"] = d
            division_lookup[f"{std_name}{div_name}"] = d

        subjects = Subject.objects.filter(school=school, is_active=True)
        subject_lookup: Dict[str, Subject] = {}
        for s in subjects:
            subject_lookup[s.name.lower().strip()] = s
            if s.code:
                subject_lookup[s.code.lower().strip()] = s

        faculties = Faculty.objects.filter(school=school, is_active=True)
        faculty_lookup: Dict[str, Faculty] = {}
        for f in faculties:
            faculty_lookup[f.full_name.lower().strip()] = f
            if f.employee_code:
                faculty_lookup[f.employee_code.lower().strip()] = f
            if f.email:
                faculty_lookup[f.email.lower().strip()] = f

        # In-memory batch tracking to catch conflicts within the same uploaded Excel
        batch_slots: Dict[Tuple[int, int, int], Dict[str, Any]] = {}  # (division_id, day, period) -> slot data
        batch_faculty_slots: Dict[Tuple[int, int, int], Dict[str, Any]] = {}  # (faculty_id, day, period) -> slot data

        row_errors: List[str] = []
        valid_records: List[Dict[str, Any]] = []

        total_processed = 0

        # 3. Process Data Rows (Row 2 onwards)
        for row_idx in range(2, ws.max_row + 1):
            row_cells = ws[row_idx]
            # Check if row is completely empty
            if all(cell.value is None or str(cell.value).strip() == '' for cell in row_cells):
                continue

            total_processed += 1

            # Extract raw values
            raw_class = str(row_cells[col_map['class']].value or '').strip()
            raw_day = str(row_cells[col_map['day']].value or '').strip().lower()
            raw_period = str(row_cells[col_map['period']].value or '').strip()
            raw_subject = str(row_cells[col_map['subject']].value or '').strip()
            raw_faculty = str(row_cells[col_map['faculty']].value or '').strip() if 'faculty' in col_map else ''
            raw_start = row_cells[col_map['start_time']].value if 'start_time' in col_map else None
            raw_end = row_cells[col_map['end_time']].value if 'end_time' in col_map else None

            # 3.1 Validate Class / Division
            clean_class_key = re.sub(r'\s+', ' ', raw_class.lower())
            division = division_lookup.get(clean_class_key)
            if not division:
                # Try relaxed matching
                clean_condensed = re.sub(r'[^a-z0-9]', '', clean_class_key)
                for k, div_obj in division_lookup.items():
                    if re.sub(r'[^a-z0-9]', '', k) == clean_condensed:
                        division = div_obj
                        break

            if not division:
                row_errors.append(f"Row {row_idx}: Class '{raw_class}' not found in active standards/divisions.")
                continue

            # 3.2 Validate User Permission for this Division (Class Teacher check)
            if not TimetableService.can_manage_division_timetable(user, division, academic_year):
                row_errors.append(f"Row {row_idx}: Permission denied. You are not authorized to manage the timetable for Class '{raw_class}'.")
                continue

            # 3.3 Validate Day of Week
            day_num = TimetableService.DAY_MAPPING.get(raw_day)
            if not day_num:
                row_errors.append(f"Row {row_idx}: Invalid day '{raw_day}'. Valid days are Monday to Saturday.")
                continue

            # 3.4 Validate Period Number
            try:
                period_num = int(raw_period)
                if not (1 <= period_num <= 12):
                    raise ValueError
            except ValueError:
                row_errors.append(f"Row {row_idx}: Invalid period number '{raw_period}'. Must be an integer between 1 and 12.")
                continue

            # 3.5 Validate Subject
            clean_sub_key = raw_subject.lower()
            subject = subject_lookup.get(clean_sub_key)
            if not subject:
                row_errors.append(f"Row {row_idx}: Subject '{raw_subject}' not found in school subject master.")
                continue

            # 3.6 Validate Faculty (Optional)
            faculty = None
            if raw_faculty:
                clean_fac_key = raw_faculty.lower()
                faculty = faculty_lookup.get(clean_fac_key)
                if not faculty:
                    row_errors.append(f"Row {row_idx}: Teacher/Faculty '{raw_faculty}' not found in active school faculty.")
                    continue

            # 3.7 Validate Times
            start_t = TimetableService.parse_time_str(raw_start)
            end_t = TimetableService.parse_time_str(raw_end)

            if raw_start and not start_t:
                row_errors.append(f"Row {row_idx}: Invalid start time format '{raw_start}'. Use HH:MM format (e.g. 08:30).")
                continue
            if raw_end and not end_t:
                row_errors.append(f"Row {row_idx}: Invalid end time format '{raw_end}'. Use HH:MM format (e.g. 09:15).")
                continue

            # 3.8 Validate Conflicts against DB
            conflicts = TimetableService.validate_slot_conflicts(
                school=school,
                academic_year=academic_year,
                division=division,
                day_of_week=day_num,
                period_number=period_num,
                faculty=faculty,
                start_time=start_t,
                end_time=end_t,
            )

            # 3.9 Validate Conflicts within Current Batch
            class_slot_key = (division.pk, day_num, period_num)
            if class_slot_key in batch_slots:
                conflicts.append(
                    f"Duplicate class period in upload: Class '{division.standard.name}-{division.name}' already has Period {period_num} defined on {raw_day.capitalize()} in this file."
                )

            if faculty:
                fac_slot_key = (faculty.pk, day_num, period_num)
                if fac_slot_key in batch_faculty_slots:
                    prev_class = batch_faculty_slots[fac_slot_key]['class_name']
                    conflicts.append(
                        f"Faculty double-booking in upload: Teacher '{faculty.full_name}' is assigned to both '{prev_class}' and '{division.standard.name}-{division.name}' on {raw_day.capitalize()} Period {period_num}."
                    )

            if conflicts:
                for c in conflicts:
                    row_errors.append(f"Row {row_idx}: {c}")
                continue

            # Register in batch trackers
            batch_slots[class_slot_key] = {'class_name': f"{division.standard.name}-{division.name}"}
            if faculty:
                batch_faculty_slots[(faculty.pk, day_num, period_num)] = {'class_name': f"{division.standard.name}-{division.name}"}

            valid_records.append({
                'division': division,
                'day_of_week': day_num,
                'period_number': period_num,
                'subject': subject,
                'faculty': faculty,
                'start_time': start_t,
                'end_time': end_t,
            })

        # 4. Save Valid Records if no fatal failures
        successful_count = 0
        if valid_records:
            with transaction.atomic():
                for rec in valid_records:
                    ClassTimetable.objects.update_or_create(
                        school=school,
                        academic_year=academic_year,
                        division=rec['division'],
                        day_of_week=rec['day_of_week'],
                        period_number=rec['period_number'],
                        defaults={
                            'subject': rec['subject'],
                            'faculty': rec['faculty'],
                            'start_time': rec['start_time'],
                            'end_time': rec['end_time'],
                        }
                    )
                    successful_count += 1

        return {
            'total_processed': total_processed,
            'successful': successful_count,
            'failed': len(row_errors),
            'errors': row_errors,
        }
