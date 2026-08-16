import io
import openpyxl
from django.db import transaction
from django.core.exceptions import ValidationError

from apps.faculty.models import Faculty
from apps.leaves.models import LeaveAllocation, LeaveType


class LeaveAllocationService:
    """
    Service layer handling Excel-based bulk leave allocation management.
    """

    @classmethod
    def generate_excel_template(cls, school):
        """
        Dynamically generates a sample leave allocation Excel template.
        Pre-populates existing faculty list for convenience.

        Returns:
            bytes: The Excel file contents as a binary stream.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leave Allocations"

        # Headers matching requirements exactly
        headers = ["Faculty ID", "Faculty Name", "Casual Leave", "Sick Leave", "Paid Leave"]
        ws.append(headers)

        # Style headers to look premium
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(
                start_color="0066CC", end_color="0066CC", fill_type="solid"
            )  # Apple Action Blue

        # Pre-populate active faculty members
        faculty_members = Faculty.objects.filter(school=school, is_active=True).order_by('employee_code')
        for fac in faculty_members:
            # Query existing allocations
            allocs = {
                a.leave_type: a.allocated 
                for a in LeaveAllocation.objects.filter(school=school, faculty=fac)
            }
            ws.append([
                fac.employee_code,
                fac.full_name,
                allocs.get(LeaveType.CASUAL, 0),
                allocs.get(LeaveType.SICK, 0),
                allocs.get(LeaveType.PAID, 0)
            ])

        # Adjust column widths dynamically
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

    @classmethod
    @transaction.atomic
    def import_leave_allocations(cls, school, file_object):
        """
        Parses and imports leave allocations from an uploaded Excel file.
        Uses transactional atomic boundaries to guarantee atomic rollback on errors.

        Args:
            school: The current School tenant.
            file_object: The uploaded Excel file object.

        Raises:
            ValidationError: If headers are missing, or row-level validation errors occur.
        """
        try:
            wb = openpyxl.load_workbook(file_object, data_only=True)
            ws = wb.active
        except Exception as e:
            raise ValidationError("Unable to read the Excel file. Please upload a valid .xlsx file.")

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValidationError("The uploaded Excel file is empty.")

        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        required_headers = ["faculty id", "casual leave", "sick leave", "paid leave"]

        # Validate headers existence
        for rh in required_headers:
            if rh not in headers:
                raise ValidationError(
                    f"Invalid template format. The sheet is missing the required '{rh.title()}' column."
                )

        # Get header indexes
        fac_idx = headers.index("faculty id")
        casual_idx = headers.index("casual leave")
        sick_idx = headers.index("sick leave")
        paid_idx = headers.index("paid leave")

        row_errors = []
        valid_rows_to_save = []
        seen_faculty_codes = set()

        for row_num, row_data in enumerate(rows[1:], start=2):
            # Skip completely empty rows
            if not any(cell_val is not None for cell_val in row_data):
                continue

            faculty_code = str(row_data[fac_idx]).strip() if row_data[fac_idx] is not None else ""
            if not faculty_code:
                row_errors.append(f"Row {row_num}: Faculty ID is required.")
                continue

            # Check duplicates within the file itself
            if faculty_code.lower() in seen_faculty_codes:
                row_errors.append(f"Row {row_num}: Duplicate entry for Faculty ID '{faculty_code}' in file.")
                continue
            seen_faculty_codes.add(faculty_code.lower())

            # Verify faculty exists in current school tenant
            faculty = Faculty.objects.filter(school=school, employee_code=faculty_code).first()
            if not faculty:
                row_errors.append(f"Row {row_num}: Faculty ID '{faculty_code}' does not exist in your school.")
                continue

            # Validate numeric leave balances
            try:
                casual = cls._parse_leave_value(row_data[casual_idx])
            except ValidationError as ve:
                row_errors.append(f"Row {row_num}: Casual Leave - {ve.message}")
                casual = None

            try:
                sick = cls._parse_leave_value(row_data[sick_idx])
            except ValidationError as ve:
                row_errors.append(f"Row {row_num}: Sick Leave - {ve.message}")
                sick = None

            try:
                paid = cls._parse_leave_value(row_data[paid_idx])
            except ValidationError as ve:
                row_errors.append(f"Row {row_num}: Paid Leave - {ve.message}")
                paid = None

            if casual is not None and sick is not None and paid is not None:
                valid_rows_to_save.append((faculty, casual, sick, paid))

        if row_errors:
            # Raise ValidationError detailing all row-level issues (this rolls back the transaction)
            raise ValidationError(row_errors)

        # Apply allocations safely
        success_count = 0
        for faculty, casual, sick, paid in valid_rows_to_save:
            for l_type, val in [
                (LeaveType.CASUAL, casual),
                (LeaveType.SICK, sick),
                (LeaveType.PAID, paid)
            ]:
                LeaveAllocation.objects.update_or_create(
                    school=school,
                    faculty=faculty,
                    leave_type=l_type,
                    defaults={'allocated': val}
                )
            success_count += 1

        return success_count

    @staticmethod
    def _parse_leave_value(val):
        """Helper to validate that a cell value represents a non-negative integer."""
        if val is None or str(val).strip() == "":
            return 0
        try:
            # Parse numbers (handles floats like 12.0 in Excel)
            num = float(val)
            if not num.is_integer():
                raise ValueError()
            int_val = int(num)
            if int_val < 0:
                raise ValueError()
            return int_val
        except (ValueError, TypeError):
            raise ValidationError("Value must be a positive/non-negative integer.")
