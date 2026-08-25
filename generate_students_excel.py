"""
Script to generate 500 realistic student test records in an Excel (.xlsx) file
matching the StudentERP1 onboarding and bulk import specifications.
"""
import random
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FIRST_NAMES_MALE = [
    "Aarav", "Vivaan", "Aditya", "Vihaan", "Arjun", "Reyansh", "Muhammad", "Sai", "Ayaan", "Krishna",
    "Ishaan", "Shaurya", "Atharv", "Advik", "Pranav", "Advaith", "Aaryan", "Dhruv", "Kabir", "Rishi",
    "Kian", "Darsh", "Rudra", "Devansh", "Samarth", "Shivansh", "Ansh", "Parth", "Yuvraj", "Ayush",
    "Veer", "Karan", "Neil", "Dev", "Siddharth", "Aayush", "Ranveer", "Daksh", "Tanmay", "Aarush",
    "Manan", "Hridaan", "Rohan", "Jayden", "Rohan", "Nirvaan", "Shlok", "Hardik", "Kavish", "Tanish",
    "Lakshya", "Chirag", "Madhav", "Kavya", "Vansh", "Vedant", "Reyan", "Yash", "Tejas", "Om"
]

FIRST_NAMES_FEMALE = [
    "Aanya", "Aadhya", "Saanvi", "Ananya", "Pari", "Diya", "Anika", "Navya", "Angel", "Myra",
    "Avani", "Sara", "Ira", "Riya", "Ahana", "Anvi", "Prisha", "Isha", "Kavya", "Kiara",
    "Meera", "Siya", "Tara", "Riddhi", "Siddhi", "Anaya", "Shanaya", "Trisha", "Vanya", "Mahi",
    "Sneha", "Pooja", "Neha", "Tanvi", "Aditi", "Bhavya", "Khushi", "Jiya", "Samaira", "Nisha",
    "Avni", "Kashvi", "Zoya", "Ruhi", "Nitya", "Kritika", "Shruti", "Swara", "Anoushka", "Tia",
    "Palak", "Radhika", "Gauri", "Ishita", "Anjali", "Muskan", "Divya", "Shreya", "Payal", "Vidhi"
]

LAST_NAMES = [
    "Patel", "Shah", "Sharma", "Verma", "Mehta", "Joshi", "Desai", "Gupta", "Singh", "Kumar",
    "Trivedi", "Pandya", "Dave", "Chauhan", "Rathore", "Bhatt", "Patolia", "Vyas", "Solanki", "Gohil",
    "Kapoor", "Malhotra", "Agarwal", "Bansal", "Reddy", "Nair", "Iyer", "Menon", "Pillai", "Choudhury",
    "Mukherjee", "Chatterjee", "Banerjee", "Das", "Ghosh", "Roy", "Sen", "Bose", "Dutta", "Dey",
    "Kulkarni", "Deshmukh", "Patil", "Pawar", "Shinde", "Jadhav", "More", "Gaekwad", "Bhide", "Apte"
]

BLOOD_GROUPS = ['A+', 'A-', 'B+', 'B-', 'O+', 'O-', 'AB+', 'AB-']

STREETS = [
    "MG Road", "Ring Road", "CG Road", "SG Highway", "Station Road", "Gandhi Nagar", "Nehru Street",
    "Subhash Marg", "Paldi Cross Road", "Satellite Area", "Vastrapur Lake Road", "Navrangpura",
    "Bodakdev High St", "Ashram Road", "Ellis Bridge Colony", "Maninagar Society", "Gota Main Road",
    "Thaltej Crescent", "Science City Road", "Prahlad Nagar Road", "Bopal Main Road", "South Bopal Avenue"
]

CITIES = ["Ahmedabad", "Surat", "Vadodara", "Rajkot", "Gandhinagar"]

def generate_student_data(total_students=500):
    records = []
    
    # 10 standards: Grade 1 to Grade 10
    # 2 divisions per standard: A and B
    # 25 students per division = 50 per grade * 10 grades = 500 students
    standards = [f"Grade {i}" for i in range(1, 11)]
    divisions = ["A", "B"]
    
    students_per_div = total_students // (len(standards) * len(divisions))
    gr_counter = 1001
    
    # Base birth year for Grade 1 is roughly 2019 (6-7 years old in 2026)
    # Grade 10 is roughly 2010 (15-16 years old in 2026)
    grade_base_years = {
        f"Grade {i}": 2020 - i for i in range(1, 11)
    }

    used_phones = set()

    for std in standards:
        birth_year = grade_base_years[std]
        for div in divisions:
            for roll_no in range(1, students_per_div + 1):
                # Gender
                is_male = (roll_no % 2 == 1)
                gender = "Male" if is_male else "Female"
                
                first_name = random.choice(FIRST_NAMES_MALE) if is_male else random.choice(FIRST_NAMES_FEMALE)
                last_name = random.choice(LAST_NAMES)
                
                # Birthday within birth year
                birth_month = random.randint(1, 12)
                birth_day = random.randint(1, 28)
                dob = f"{birth_year:04d}-{birth_month:02d}-{birth_day:02d}"
                
                blood_group = random.choice(BLOOD_GROUPS)
                
                # Guardian Details
                father_name = random.choice(FIRST_NAMES_MALE)
                guardian_name = f"{father_name} {last_name}"
                
                # Phone generator (ensure unique valid 10-digit starting with 98, 97, 96, 99, etc.)
                while True:
                    prefix = random.choice(["98", "99", "97", "96", "94", "93", "91", "90"])
                    suffix = f"{random.randint(10000000, 99999999)}"[:8]
                    phone = prefix + suffix
                    if phone not in used_phones:
                        used_phones.add(phone)
                        break
                
                # Emergency contact
                while True:
                    prefix = random.choice(["98", "99", "97", "96", "94", "93", "91", "90"])
                    suffix = f"{random.randint(10000000, 99999999)}"[:8]
                    emer_phone = prefix + suffix
                    if emer_phone != phone and emer_phone not in used_phones:
                        used_phones.add(emer_phone)
                        break

                clean_first = first_name.lower().replace(" ", "")
                clean_last = last_name.lower().replace(" ", "")
                parent_email = f"{clean_first}.{clean_last}.parent@example.com"
                
                house_no = random.randint(1, 999)
                street = random.choice(STREETS)
                city = random.choice(CITIES)
                pincode = random.randint(380001, 380060)
                address = f"{house_no}, {street}, {city} - {pincode}"
                
                gr_number = f"GR-{gr_counter}"
                gr_counter += 1
                
                record = {
                    "GR Number": gr_number,
                    "First Name": first_name,
                    "Last Name": last_name,
                    "Standard Name": std,
                    "Division Name": div,
                    "Roll Number": roll_no,
                    "Gender": gender,
                    "Date of Birth": dob,
                    "Blood Group": blood_group,
                    "Guardian Name": guardian_name,
                    "Parent Phone": phone,
                    "Parent Email": parent_email,
                    "Emergency Contact": emer_phone,
                    "Address": address,
                }
                records.append(record)
                
    return records


def create_styled_excel(records, output_filename="students_500_test_data.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Step 4 Sample Import"
    
    headers = [
        "GR Number",
        "First Name",
        "Last Name",
        "Standard Name",
        "Division Name",
        "Roll Number",
        "Gender",
        "Date of Birth",
        "Blood Group",
        "Guardian Name",
        "Parent Phone",
        "Parent Email",
        "Emergency Contact",
        "Address",
    ]
    
    # Styles matching Apple Design / Onboarding Service
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(name="Segoe UI", size=10, color="1D1D1F")
    alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="E5E7EB")
    cell_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # Write Header
    ws.append(headers)
    ws.row_dimensions[1].height = 28
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border

    # Center-aligned columns
    center_cols = {"GR Number", "Standard Name", "Division Name", "Roll Number", "Gender", "Date of Birth", "Blood Group", "Parent Phone", "Emergency Contact"}

    # Write 500 Data Rows
    for row_idx, rec in enumerate(records, start=2):
        row_data = [rec[h] for h in headers]
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 20
        
        row_fill = alt_fill if (row_idx % 2 == 0) else white_fill
        
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = cell_border
            if h in center_cols:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Auto-adjust column widths with generous padding
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Freeze panes below header
    ws.freeze_panes = "A2"
    
    wb.save(output_filename)
    print(f"Successfully generated {len(records)} student records in '{output_filename}'")


if __name__ == "__main__":
    records = generate_student_data(500)
    create_styled_excel(records, "students_500_test_data.xlsx")
    # Also create step_4_sample_template.xlsx as a companion standard template
    create_styled_excel(records, "step_4_sample_template.xlsx")
