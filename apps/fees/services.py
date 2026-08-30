"""
Simplified School Fees Management Services.

Provides:
  - FeeExcelService: Generates downloadable Excel template and imports class-wise fee structures.
  - FeeService: Direct payment recording, receipt numbering, student fee summaries, and metrics.
"""
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional, Tuple, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db import models, transaction
from django.utils import timezone
from django.core.exceptions import ValidationError

from apps.accounts.models import User
from apps.academics.models import AcademicYear, Standard, Division
from apps.students.models import Student
from apps.fees.models import FeeCategory, FeeStructure, StudentFee, FeePayment
from apps.notifications.models import InAppNotification


class FeeExcelService:
    """
    Service for generating downloadable Fee Structure Excel template and processing bulk uploads.
    """

    FREQUENCY_MAP = {
        'monthly': FeeStructure.Frequency.MONTHLY,
        'month': FeeStructure.Frequency.MONTHLY,
        '12': FeeStructure.Frequency.MONTHLY,
        'quarterly': FeeStructure.Frequency.QUARTERLY,
        'quarter': FeeStructure.Frequency.QUARTERLY,
        '4': FeeStructure.Frequency.QUARTERLY,
        'half-yearly': FeeStructure.Frequency.HALF_YEARLY,
        'halfyearly': FeeStructure.Frequency.HALF_YEARLY,
        'half yearly': FeeStructure.Frequency.HALF_YEARLY,
        'semi-annual': FeeStructure.Frequency.HALF_YEARLY,
        '2': FeeStructure.Frequency.HALF_YEARLY,
        'full-year': FeeStructure.Frequency.FULL_YEAR,
        'fullyear': FeeStructure.Frequency.FULL_YEAR,
        'full year': FeeStructure.Frequency.FULL_YEAR,
        'annual': FeeStructure.Frequency.FULL_YEAR,
        'yearly': FeeStructure.Frequency.FULL_YEAR,
        '1': FeeStructure.Frequency.FULL_YEAR,
    }

    @classmethod
    def generate_sample_template(cls, school, academic_year: AcademicYear) -> bytes:
        """
        Creates an Apple-styled Excel workbook containing:
        1. 'Fee Structure' sheet with Class, Total Fee, Payment Frequency.
        2. 'Available Classes' reference sheet with existing classes in the school.
        """
        wb = openpyxl.Workbook()

        # Sheet 1: Main Upload Sheet
        ws = wb.active
        ws.title = "Fee Structure"

        headers = ["Class", "Total Fee", "Payment Frequency"]
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0'),
        )

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
            ws.row_dimensions[1].height = 28

        # Get existing standards for realistic samples
        standards = Standard.objects.filter(school=school, is_active=True).order_by('order_index')
        sample_rows = []
        if standards.exists():
            frequencies = ["Monthly", "Quarterly", "Half-Yearly", "Full-Year"]
            for idx, std in enumerate(standards[:6], start=1):
                sample_amt = 25000 + (idx * 3000)
                sample_freq = frequencies[(idx - 1) % len(frequencies)]
                sample_rows.append([std.name, sample_amt, sample_freq])
        else:
            sample_rows = [
                ["1", 25000, "Monthly"],
                ["2", 28000, "Monthly"],
                ["3", 30000, "Quarterly"],
                ["4", 35000, "Half-Yearly"],
                ["5", 40000, "Full-Year"],
                ["10-A", 50000, "Quarterly"],
            ]

        sample_font = Font(name="Arial", size=10)
        for row in sample_rows:
            ws.append(row)
            row_idx = ws.max_row
            ws.row_dimensions[row_idx].height = 22
            for col_num in range(1, len(row) + 1):
                c = ws.cell(row=row_idx, column=col_num)
                c.font = sample_font
                c.border = border
                if col_num == 2:
                    c.alignment = right_align
                    c.number_format = '#,##0.00'
                elif col_num == 3:
                    c.alignment = center_align
                else:
                    c.alignment = left_align

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 20)

        # Sheet 2: Reference Available Classes
        ws_ref = wb.create_sheet(title="Available Classes")
        ref_header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")

        ref_headers = ["Class / Standard Name", "Grade Order", "Division (Optional)"]
        ws_ref.append(ref_headers)
        ws_ref.row_dimensions[1].height = 26
        for col_num in range(1, 4):
            c = ws_ref.cell(row=1, column=col_num)
            c.font = header_font
            c.fill = ref_header_fill
            c.alignment = center_align
            c.border = border

        divisions = Division.objects.filter(school=school, is_active=True).select_related('standard').order_by('standard__order_index', 'name')
        if divisions.exists():
            for div in divisions:
                ws_ref.append([f"{div.standard.name} - {div.name}", div.standard.order_index, div.name])
                r_idx = ws_ref.max_row
                ws_ref.row_dimensions[r_idx].height = 20
                for col_idx in range(1, 4):
                    cell = ws_ref.cell(row=r_idx, column=col_idx)
                    cell.font = sample_font
                    cell.border = border
        else:
            for std in standards:
                ws_ref.append([std.name, std.order_index, "All Divisions"])
                r_idx = ws_ref.max_row
                ws_ref.row_dimensions[r_idx].height = 20
                for col_idx in range(1, 4):
                    cell = ws_ref.cell(row=r_idx, column=col_idx)
                    cell.font = sample_font
                    cell.border = border

        for col in ws_ref.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_ref.column_dimensions[col_letter].width = max(max_len + 4, 22)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @classmethod
    def import_fee_structure_excel(
        cls,
        school,
        academic_year: AcademicYear,
        file_obj,
        user: User,
    ) -> Dict[str, Any]:
        """
        Parses and validates the uploaded Fee Structure Excel file.
        Creates/updates FeeStructure per class and automatically syncs StudentFee records.
        """
        filename = getattr(file_obj, 'name', 'upload.xlsx').lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            return {
                'total_processed': 0,
                'successful': 0,
                'failed': 0,
                'errors': ["Invalid file type. Please upload a valid Microsoft Excel (.xlsx or .xls) file."],
            }

        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
        except Exception as e:
            return {
                'total_processed': 0,
                'successful': 0,
                'failed': 0,
                'errors': [f"Could not open Excel file: {str(e)}"],
            }

        # 1. Map Header Columns
        header_row = [str(cell.value or '').strip().lower() for cell in ws[1]]
        col_map = {}
        for idx, col_name in enumerate(header_row):
            clean_name = re.sub(r'[^a-z0-9]', '', col_name)
            if 'class' in clean_name or 'standard' in clean_name or 'grade' in clean_name:
                col_map['class'] = idx
            elif 'fee' in clean_name or 'amount' in clean_name or 'total' in clean_name:
                col_map['amount'] = idx
            elif 'frequency' in clean_name or 'freq' in clean_name or 'payment' in clean_name:
                col_map['frequency'] = idx

        required_cols = ['class', 'amount', 'frequency']
        missing = [c.capitalize() for c in required_cols if c not in col_map]
        if missing:
            return {
                'total_processed': 0,
                'successful': 0,
                'failed': 0,
                'errors': [f"Missing required columns in Excel sheet: {', '.join(missing)}. Please use the downloadable template."],
            }

        # 2. Build In-Memory Class Lookups for the School Tenant
        standards = Standard.objects.filter(school=school, is_active=True)
        standard_lookup: Dict[str, Standard] = {}
        for std in standards:
            name_clean = std.name.lower().strip()
            standard_lookup[name_clean] = std
            standard_lookup[str(std.order_index)] = std
            standard_lookup[f"class {std.order_index}"] = std
            standard_lookup[f"std {std.order_index}"] = std
            standard_lookup[f"standard {std.order_index}"] = std
            standard_lookup[f"grade {std.order_index}"] = std

        divisions = Division.objects.filter(school=school, is_active=True).select_related('standard')
        division_lookup: Dict[str, Division] = {}
        for div in divisions:
            d_name = div.name.lower().strip()
            std_name = div.standard.name.lower().strip()
            division_lookup[f"{std_name}-{d_name}"] = div
            division_lookup[f"{std_name} {d_name}"] = div
            division_lookup[f"{div.standard.order_index}-{d_name}"] = div
            division_lookup[f"{div.standard.order_index}{d_name}"] = div
            division_lookup[f"{std_name}{d_name}"] = div

        default_category = FeeService.ensure_default_category(school)

        row_errors: List[str] = []
        valid_records: List[Dict[str, Any]] = []
        seen_classes = set()
        total_processed = 0

        # 3. Process Data Rows (Row 2 onwards)
        for row_idx in range(2, ws.max_row + 1):
            row_cells = ws[row_idx]
            if all(cell.value is None or str(cell.value).strip() == '' for cell in row_cells):
                continue

            total_processed += 1

            raw_class = str(row_cells[col_map['class']].value or '').strip()
            raw_amount = str(row_cells[col_map['amount']].value or '').strip()
            raw_freq_display = str(row_cells[col_map['frequency']].value or '').strip()
            raw_freq = raw_freq_display.lower()

            # 3.1 Validate Class
            clean_class_key = re.sub(r'\s+', ' ', raw_class.lower())
            target_standard = standard_lookup.get(clean_class_key)
            target_division = division_lookup.get(clean_class_key)

            if not target_standard and not target_division:
                clean_condensed = re.sub(r'[^a-z0-9]', '', clean_class_key)
                for k, std_obj in standard_lookup.items():
                    if re.sub(r'[^a-z0-9]', '', k) == clean_condensed:
                        target_standard = std_obj
                        break
                if not target_standard:
                    for k, div_obj in division_lookup.items():
                        if re.sub(r'[^a-z0-9]', '', k) == clean_condensed:
                            target_division = div_obj
                            break

            if not target_standard and not target_division:
                row_errors.append(f"Row {row_idx}: Class '{raw_class}' not found in active standards or divisions.")
                continue

            # 3.2 Duplicate Class in File
            class_id_key = f"div_{target_division.pk}" if target_division else f"std_{target_standard.pk}"
            if class_id_key in seen_classes:
                row_errors.append(f"Row {row_idx}: Duplicate entry for Class '{raw_class}'. Each class should only have one fee structure.")
                continue
            seen_classes.add(class_id_key)

            # 3.3 Validate Fee Amount
            try:
                # Remove currency symbols or commas if present
                clean_amt_str = re.sub(r'[^\d.-]', '', raw_amount)
                amount = Decimal(clean_amt_str)
                if amount <= Decimal('0.00'):
                    raise ValueError
            except (InvalidOperation, ValueError):
                row_errors.append(f"Row {row_idx}: Invalid fee amount '{raw_amount}'. Must be a positive number.")
                continue

            # 3.4 Validate Payment Frequency
            clean_freq_key = re.sub(r'[^a-z0-9]', '', raw_freq)
            freq_value = cls.FREQUENCY_MAP.get(raw_freq) or cls.FREQUENCY_MAP.get(clean_freq_key)
            if not freq_value:
                row_errors.append(f"Row {row_idx}: Invalid payment frequency '{raw_freq_display}'. Allowed: Monthly, Quarterly, Half-Yearly, Full-Year.")
                continue

            valid_records.append({
                'standard': target_standard or (target_division.standard if target_division else None),
                'division': target_division,
                'amount': amount,
                'frequency': freq_value,
                'class_display': raw_class,
            })

        # 4. Save Valid Fee Structures & Automatically Sync Students
        successful_count = 0
        if valid_records:
            with transaction.atomic():
                for rec in valid_records:
                    std = rec['standard']
                    div = rec['division']
                    amt = rec['amount']
                    freq = rec['frequency']

                    name = f"{std.name} Annual Fee" if std else f"{div.standard.name}-{div.name} Annual Fee"

                    # Find or create FeeStructure
                    structure, created = FeeStructure.objects.update_or_create(
                        school=school,
                        academic_year=academic_year,
                        standard=std,
                        division=div,
                        defaults={
                            'name': name,
                            'fee_category': default_category,
                            'amount': amt,
                            'payment_frequency': freq,
                            'is_active': True,
                        }
                    )

                    # Auto-sync/propagate to all students in that class
                    FeeService.sync_students_for_structure(
                        school=school,
                        academic_year=academic_year,
                        fee_structure=structure,
                    )
                    successful_count += 1

        return {
            'total_processed': total_processed,
            'successful': successful_count,
            'failed': len(row_errors),
            'errors': row_errors,
        }


class FeeService:
    """
    Simplified service for recording payments, compiling student fee summaries,
    and managing school fee metrics.
    """

    @classmethod
    def ensure_default_category(cls, school) -> FeeCategory:
        """
        Retrieves or creates a standard default FeeCategory for the school.
        """
        category, _ = FeeCategory.objects.get_or_create(
            school=school,
            name="Tuition & Academic Fee",
            defaults={
                'code': 'TUITION',
                'description': 'Standard school tuition and academic fee',
                'is_active': True,
            }
        )
        return category

    @classmethod
    def generate_next_receipt_number(cls, school) -> str:
        """
        Generates a unique sequential receipt number for the school:
        Format: REC-{SUBDOMAIN}-{YEAR}-{SEQUENCE:05d}
        """
        current_year = timezone.localdate().year
        prefix = f"REC-{school.subdomain.upper()}-{current_year}-"

        last_payment = FeePayment.objects.filter(
            school=school,
            receipt_number__startswith=prefix,
        ).order_by('-id').first()

        if last_payment and last_payment.receipt_number:
            try:
                last_seq = int(last_payment.receipt_number.split('-')[-1])
                next_seq = last_seq + 1
            except (ValueError, IndexError):
                next_seq = 1
        else:
            next_seq = 1

        return f"{prefix}{next_seq:05d}"

    @classmethod
    def sync_students_for_structure(cls, school, academic_year: AcademicYear, fee_structure: FeeStructure) -> int:
        """
        Automatically creates or updates StudentFee assignments for all active students in the target standard/division.
        Preserves all existing payments and adjusts net payable amounts cleanly.
        """
        students_qs = Student.objects.filter(
            school=school,
            academic_year=academic_year,
            is_active=True,
        )
        if fee_structure.division:
            students_qs = students_qs.filter(division=fee_structure.division)
        elif fee_structure.standard:
            students_qs = students_qs.filter(standard=fee_structure.standard)

        count = 0
        for student in students_qs:
            StudentFee.objects.update_or_create(
                school=school,
                student=student,
                academic_year=academic_year,
                defaults={
                    'fee_structure': fee_structure,
                    'is_active': True,
                }
            )
            count += 1
        return count

    @classmethod
    @transaction.atomic
    def record_payment(
        cls,
        school,
        student: Student,
        amount: Decimal,
        payment_date: Optional[date] = None,
        payment_method: str = FeePayment.PaymentMethod.CASH,
        transaction_reference: str = '',
        recorded_by: Optional[User] = None,
    ) -> Tuple[FeePayment, Optional[StudentFee]]:
        """
        Records an approved fee payment for a student and generates an official receipt.
        """
        if amount <= Decimal('0.00'):
            raise ValidationError("Payment amount must be greater than zero.")

        if payment_date is None:
            payment_date = timezone.localdate()

        academic_year = student.academic_year
        if not academic_year:
            academic_year = AcademicYear.objects.filter(school=school, is_current=True).first()

        # Ensure student has an assigned StudentFee record
        student_fee = StudentFee.objects.filter(
            school=school,
            student=student,
            academic_year=academic_year,
            is_active=True,
        ).first()

        if not student_fee:
            # Try to auto-link matching FeeStructure for student's standard
            structure = FeeStructure.objects.filter(
                school=school,
                academic_year=academic_year,
                standard=student.standard,
                is_active=True,
            ).first()
            if not structure:
                # Fallback to any general active structure
                structure = FeeStructure.objects.filter(
                    school=school,
                    academic_year=academic_year,
                    standard__isnull=True,
                    is_active=True,
                ).first()

            if structure:
                student_fee = StudentFee.objects.create(
                    school=school,
                    student=student,
                    academic_year=academic_year,
                    fee_structure=structure,
                    is_active=True,
                )

        receipt_number = cls.generate_next_receipt_number(school)

        payment = FeePayment.objects.create(
            school=school,
            student=student,
            academic_year=academic_year,
            amount=amount,
            payment_date=payment_date,
            payment_method=payment_method,
            transaction_reference=transaction_reference.strip(),
            receipt_number=receipt_number,
            recorded_by=recorded_by,
            status=FeePayment.Status.SUCCESS,
        )

        # Dispatch in-app notification if student user exists
        if student.user:
            try:
                InAppNotification.objects.create(
                    school=school,
                    user=student.user,
                    title="Fee Payment Received",
                    message=f"Payment of ₹{amount} has been successfully recorded. Receipt #{receipt_number} is available in your portal.",
                )
            except Exception:
                pass

        return payment, student_fee

    @classmethod
    def get_student_fee_summary(cls, student: Student, academic_year: Optional[AcademicYear] = None) -> Dict[str, Any]:
        """
        Compiles the streamlined student fee summary: Total Fee, Paid, Remaining, Frequency, Status, and Payment history.
        """
        school = student.school
        if academic_year is None:
            academic_year = student.academic_year or AcademicYear.objects.filter(school=school, is_current=True).first()

        student_fee = StudentFee.objects.filter(
            school=school,
            student=student,
            academic_year=academic_year,
            is_active=True,
        ).select_related('fee_structure').first()

        # If not explicitly created, check standard's fee structure
        if not student_fee and academic_year and student.standard:
            structure = FeeStructure.objects.filter(
                school=school,
                academic_year=academic_year,
                standard=student.standard,
                is_active=True,
            ).first()
            if structure:
                student_fee = StudentFee(
                    school=school,
                    student=student,
                    academic_year=academic_year,
                    fee_structure=structure,
                )

        payments = list(
            FeePayment.objects.filter(
                school=school,
                student=student,
                academic_year=academic_year,
                status=FeePayment.Status.SUCCESS,
            ).order_by('-payment_date', '-id')
        )

        total_paid = sum((p.amount for p in payments), Decimal('0.00'))

        if student_fee:
            total_fee = student_fee.net_payable_amount
            frequency = student_fee.fee_structure.get_payment_frequency_display()
            fee_name = student_fee.fee_structure.name
        else:
            total_fee = Decimal('0.00')
            frequency = "Full-Year"
            fee_name = "Standard Fee"

        remaining_amount = max(Decimal('0.00'), total_fee - total_paid)

        if total_fee == Decimal('0.00') and total_paid == Decimal('0.00'):
            status = 'NOT ASSIGNED'
        elif remaining_amount == Decimal('0.00'):
            status = 'PAID'
        elif total_paid > Decimal('0.00'):
            status = 'PARTIALLY PAID'
        else:
            status = 'PENDING'

        return {
            'has_fee': (student_fee is not None or total_fee > 0),
            'student_fee': student_fee,
            'fee_structure_name': fee_name,
            'frequency': frequency,
            'total_fee': total_fee,
            'total_paid': total_paid,
            'remaining_amount': remaining_amount,
            'status': status,
            'payments': payments,
            'academic_year': academic_year,
        }

    @classmethod
    def get_school_fee_metrics(cls, school, academic_year: Optional[AcademicYear] = None) -> Dict[str, Any]:
        """
        Aggregates school-wide fee metrics for the session.
        """
        if academic_year is None:
            academic_year = AcademicYear.objects.filter(school=school, is_current=True).first()

        students = Student.objects.filter(
            school=school,
            academic_year=academic_year,
            is_active=True,
        ).select_related('standard')

        total_expected = Decimal('0.00')
        for s in students:
            summary = cls.get_student_fee_summary(student=s, academic_year=academic_year)
            total_expected += summary['total_fee']

        total_collected = FeePayment.objects.filter(
            school=school,
            academic_year=academic_year,
            status=FeePayment.Status.SUCCESS,
        ).aggregate(total=models.Sum('amount'))['total'] or Decimal('0.00')

        total_outstanding = max(Decimal('0.00'), total_expected - total_collected)
        rate_pct = int((total_collected / total_expected * 100)) if total_expected > 0 else 0

        return {
            'total_expected': total_expected,
            'total_collected': total_collected,
            'total_outstanding': total_outstanding,
            'collection_rate_pct': min(100, rate_pct),
        }
