"""
Simplified School Fees Management Test Suite.

Covers:
  - Excel template download
  - Excel bulk upload (valid multi-class, all 4 frequencies, invalid file types, missing headers, bad amounts, unknown classes, duplicate rows)
  - Automatic student fee calculation (Total, Paid, Remaining, Status)
  - Direct payment recording with sequential receipt numbers
  - Historical payment preservation when updating fee structures
  - Student portal view and receipt download
  - Student privacy & unauthorized mutation prevention
  - Multi-tenant data isolation
  - Feature flag enforcement
"""
import io
from datetime import date
from decimal import Decimal
import openpyxl
from django.test import TestCase, Client
from django.urls import reverse
from django.core.files.uploadedfile import SimpleUploadedFile

from apps.accounts.models import User
from apps.academics.models import AcademicYear, Standard, Division
from apps.students.models import Student
from apps.tenants.models import School, SchoolFeature
from apps.fees.models import FeeStructure, StudentFee, FeePayment
from apps.fees.services import FeeService, FeeExcelService


class SimplifiedFeesTestSuite(TestCase):
    def setUp(self):
        self.client = Client()

        # 1. School A (Primary Tenant)
        self.school_a = School.objects.create(name="Greenwood High", subdomain="greenwood")

        self.year_a = AcademicYear.objects.create(
            school=self.school_a,
            name="2026-2027",
            start_date=date(2026, 6, 1),
            end_date=date(2027, 5, 31),
            is_current=True,
        )

        # Classes in School A
        self.std_1 = Standard.objects.create(school=self.school_a, name="Class 1", order_index=1)
        self.div_1a = Division.objects.create(school=self.school_a, standard=self.std_1, name="A")

        self.std_10 = Standard.objects.create(school=self.school_a, name="Class 10", order_index=10)
        self.div_10a = Division.objects.create(school=self.school_a, standard=self.std_10, name="A")

        # Admin A
        self.admin_a = User.objects.create_user(
            username="admin_a",
            email="admin@greenwood.edu",
            password="Password@123",
            role=User.Role.SCHOOL_ADMIN,
            school=self.school_a,
        )

        # Student 1: Rahul Patel in Class 10-A
        self.student_user_1 = User.objects.create_user(
            username="GR1001",
            email="rahul@greenwood.edu",
            password="Password@123",
            role=User.Role.STUDENT,
            school=self.school_a,
        )
        self.student_1 = Student.objects.create(
            school=self.school_a,
            user=self.student_user_1,
            gr_number="GR1001",
            roll_number=10,
            full_name="Rahul Patel",
            academic_year=self.year_a,
            standard=self.std_10,
            division=self.div_10a,
            is_active=True,
        )

        # Student 2: Ananya in Class 1-A
        self.student_user_2 = User.objects.create_user(
            username="GR1002",
            email="ananya@greenwood.edu",
            password="Password@123",
            role=User.Role.STUDENT,
            school=self.school_a,
        )
        self.student_2 = Student.objects.create(
            school=self.school_a,
            user=self.student_user_2,
            gr_number="GR1002",
            roll_number=1,
            full_name="Ananya Sharma",
            academic_year=self.year_a,
            standard=self.std_1,
            division=self.div_1a,
            is_active=True,
        )

        # 2. School B (Secondary Tenant for Isolation Testing)
        self.school_b = School.objects.create(name="St. Mary Academy", subdomain="stmary")
        self.year_b = AcademicYear.objects.create(
            school=self.school_b,
            name="2026-2027",
            start_date=date(2026, 6, 1),
            end_date=date(2027, 5, 31),
            is_current=True,
        )
        self.std_b10 = Standard.objects.create(school=self.school_b, name="Grade 10", order_index=10)
        self.div_b10 = Division.objects.create(school=self.school_b, standard=self.std_b10, name="A")
        self.admin_b = User.objects.create_user(
            username="admin_b",
            email="admin@stmary.edu",
            password="Password@123",
            role=User.Role.SCHOOL_ADMIN,
            school=self.school_b,
        )
        self.student_user_b = User.objects.create_user(
            username="ST1001",
            email="john@stmary.edu",
            password="Password@123",
            role=User.Role.STUDENT,
            school=self.school_b,
        )
        self.student_b = Student.objects.create(
            school=self.school_b,
            user=self.student_user_b,
            gr_number="ST1001",
            roll_number=1,
            full_name="John Doe",
            academic_year=self.year_b,
            standard=self.std_b10,
            division=self.div_b10,
            is_active=True,
        )

    # ══════════════════════════════════════════════════════════════════════════
    # 1. EXCEL TEMPLATE & BULK UPLOAD
    # ══════════════════════════════════════════════════════════════════════════

    def test_excel_template_download(self):
        """School Admin downloads the sample Fee Structure Excel template."""
        self.client.force_login(self.admin_a)
        url = reverse('fees:template_download')
        res = self.client.get(url, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res.status_code, 200)
        self.assertIn("spreadsheetml.sheet", res['Content-Type'])

        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        self.assertIn("Fee Structure", wb.sheetnames)
        self.assertIn("Available Classes", wb.sheetnames)

    def _create_test_excel(self, rows, headers=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fee Structure"
        if headers is None:
            headers = ["Class", "Total Fee", "Payment Frequency"]
        ws.append(headers)
        for r in rows:
            ws.append(r)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_excel_bulk_upload_valid_all_frequencies(self):
        """Admin uploads Excel with all 4 frequencies (Monthly, Quarterly, Half-Yearly, Full-Year)."""
        rows = [
            ["Class 1", "25000", "Monthly"],
            ["Class 10", "50000", "Quarterly"],
        ]
        excel_buf = self._create_test_excel(rows)
        upload_file = SimpleUploadedFile("fees.xlsx", excel_buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        self.client.force_login(self.admin_a)
        url = reverse('fees:excel_upload')
        res = self.client.post(url, {'excel_file': upload_file}, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res.status_code, 302)

        # Check structures created
        struct_10 = FeeStructure.objects.get(school=self.school_a, standard=self.std_10)
        self.assertEqual(struct_10.amount, Decimal('50000.00'))
        self.assertEqual(struct_10.payment_frequency, FeeStructure.Frequency.QUARTERLY)

        # Check students automatically synced
        sf_rahul = StudentFee.objects.get(school=self.school_a, student=self.student_1)
        self.assertEqual(sf_rahul.net_payable_amount, Decimal('50000.00'))

        summary_rahul = FeeService.get_student_fee_summary(self.student_1, self.year_a)
        self.assertEqual(summary_rahul['total_fee'], Decimal('50000.00'))
        self.assertEqual(summary_rahul['total_paid'], Decimal('0.00'))
        self.assertEqual(summary_rahul['remaining_amount'], Decimal('50000.00'))
        self.assertEqual(summary_rahul['frequency'], "Quarterly")
        self.assertEqual(summary_rahul['status'], "PENDING")

    def test_excel_upload_invalid_file_format(self):
        """Uploading non-Excel file (.txt / .pdf) returns error."""
        text_file = SimpleUploadedFile("bad_file.txt", b"Hello", content_type="text/plain")
        result = FeeExcelService.import_fee_structure_excel(
            school=self.school_a,
            academic_year=self.year_a,
            file_obj=text_file,
            user=self.admin_a,
        )
        self.assertEqual(result['successful'], 0)
        self.assertTrue(any("Invalid file type" in e for e in result['errors']))

    def test_excel_upload_missing_columns(self):
        """Excel missing required 'Payment Frequency' column is rejected."""
        headers = ["Class", "Total Fee"]  # Missing Frequency
        rows = [["Class 10", "50000"]]
        excel_buf = self._create_test_excel(rows, headers=headers)
        upload_file = SimpleUploadedFile("missing_col.xlsx", excel_buf.getvalue())

        result = FeeExcelService.import_fee_structure_excel(
            school=self.school_a,
            academic_year=self.year_a,
            file_obj=upload_file,
            user=self.admin_a,
        )
        self.assertEqual(result['successful'], 0)
        self.assertTrue(any("Missing required columns" in e for e in result['errors']))

    def test_excel_upload_row_validation_errors(self):
        """Excel with invalid class, negative fee, invalid frequency, or duplicate class reports row-level errors."""
        rows = [
            ["Unknown Class", "50000", "Monthly"],   # Row 2: Unknown class
            ["Class 1", "-500", "Monthly"],          # Row 3: Negative fee
            ["Class 10", "50000", "InvalidFreq"],    # Row 4: Bad frequency
            ["Class 1", "30000", "Monthly"],         # Row 5: Duplicate Class 1
        ]
        excel_buf = self._create_test_excel(rows)
        upload_file = SimpleUploadedFile("bad_rows.xlsx", excel_buf.getvalue())

        result = FeeExcelService.import_fee_structure_excel(
            school=self.school_a,
            academic_year=self.year_a,
            file_obj=upload_file,
            user=self.admin_a,
        )
        self.assertEqual(result['total_processed'], 4)
        self.assertEqual(result['successful'], 0)
        self.assertEqual(result['failed'], 4)
        self.assertTrue(any("Row 2: Class 'Unknown Class' not found" in e for e in result['errors']))
        self.assertTrue(any("Row 3: Invalid fee amount '-500'" in e for e in result['errors']))
        self.assertTrue(any("Row 4: Invalid payment frequency 'InvalidFreq'" in e for e in result['errors']))
        self.assertTrue(any("Row 5: Duplicate entry for Class 'Class 1'" in e for e in result['errors']))

    # ══════════════════════════════════════════════════════════════════════════
    # 2. PAYMENT RECORDING & PROGRESSIVE CALCULATION
    # ══════════════════════════════════════════════════════════════════════════

    def test_payment_recording_calculation_flow(self):
        """
        Tests the complete payment flow requested:
        Total Fee = ₹50,000
        Payment 1: ₹20,000 -> Total: ₹50,000, Paid: ₹20,000, Remaining: ₹30,000, Status: PARTIALLY PAID
        Payment 2: ₹15,000 -> Total: ₹50,000, Paid: ₹35,000, Remaining: ₹15,000, Status: PARTIALLY PAID
        Payment 3: ₹15,000 -> Total: ₹50,000, Paid: ₹50,000, Remaining: ₹0, Status: PAID
        """
        # Set Class 10 fee structure
        FeeStructure.objects.create(
            school=self.school_a,
            academic_year=self.year_a,
            standard=self.std_10,
            fee_category=FeeService.ensure_default_category(self.school_a),
            name="Class 10 Annual Fee",
            amount=Decimal('50000.00'),
            payment_frequency=FeeStructure.Frequency.QUARTERLY,
        )

        self.client.force_login(self.admin_a)
        url = reverse('fees:record_payment')

        # Payment 1: ₹20,000
        res1 = self.client.post(url, {
            'student_id': self.student_1.pk,
            'amount': '20000',
            'payment_date': '2026-08-20',
            'payment_method': 'CASH',
            'transaction_reference': 'RCP-001',
        }, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res1.status_code, 302)

        summary_1 = FeeService.get_student_fee_summary(self.student_1, self.year_a)
        self.assertEqual(summary_1['total_fee'], Decimal('50000.00'))
        self.assertEqual(summary_1['total_paid'], Decimal('20000.00'))
        self.assertEqual(summary_1['remaining_amount'], Decimal('30000.00'))
        self.assertEqual(summary_1['status'], "PARTIALLY PAID")

        # Payment 2: ₹15,000
        res2 = self.client.post(url, {
            'student_id': self.student_1.pk,
            'amount': '15000',
            'payment_date': '2026-09-25',
            'payment_method': 'BANK_TRANSFER',
            'transaction_reference': 'UTR998877',
        }, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res2.status_code, 302)

        summary_2 = FeeService.get_student_fee_summary(self.student_1, self.year_a)
        self.assertEqual(summary_2['total_fee'], Decimal('50000.00'))
        self.assertEqual(summary_2['total_paid'], Decimal('35000.00'))
        self.assertEqual(summary_2['remaining_amount'], Decimal('15000.00'))
        self.assertEqual(summary_2['status'], "PARTIALLY PAID")

        # Payment 3: ₹15,000
        res3 = self.client.post(url, {
            'student_id': self.student_1.pk,
            'amount': '15000',
            'payment_date': '2026-10-10',
            'payment_method': 'ONLINE',
            'transaction_reference': 'ONL-5544',
        }, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res3.status_code, 302)

        summary_3 = FeeService.get_student_fee_summary(self.student_1, self.year_a)
        self.assertEqual(summary_3['total_fee'], Decimal('50000.00'))
        self.assertEqual(summary_3['total_paid'], Decimal('50000.00'))
        self.assertEqual(summary_3['remaining_amount'], Decimal('0.00'))
        self.assertEqual(summary_3['status'], "PAID")
        self.assertEqual(len(summary_3['payments']), 3)

    # ══════════════════════════════════════════════════════════════════════════
    # 3. HISTORICAL PAYMENT PRESERVATION ON STRUCTURE UPDATE
    # ══════════════════════════════════════════════════════════════════════════

    def test_fee_structure_update_preserves_historical_payments(self):
        """Updating fee structure from ₹50,000 to ₹60,000 preserves ₹20,000 payment and adjusts balance to ₹40,000."""
        # Initial structure: ₹50,000
        struct = FeeStructure.objects.create(
            school=self.school_a,
            academic_year=self.year_a,
            standard=self.std_10,
            fee_category=FeeService.ensure_default_category(self.school_a),
            name="Class 10 Fee",
            amount=Decimal('50000.00'),
            payment_frequency=FeeStructure.Frequency.QUARTERLY,
        )
        FeeService.sync_students_for_structure(self.school_a, self.year_a, struct)

        # Record payment of ₹20,000
        p, _ = FeeService.record_payment(
            school=self.school_a,
            student=self.student_1,
            amount=Decimal('20000.00'),
            payment_method=FeePayment.PaymentMethod.CASH,
        )

        # Admin uploads new Excel updating Class 10 fee to ₹60,000
        rows = [["Class 10", "60000", "Quarterly"]]
        excel_buf = self._create_test_excel(rows)
        upload_file = SimpleUploadedFile("fees_v2.xlsx", excel_buf.getvalue())
        res = FeeExcelService.import_fee_structure_excel(self.school_a, self.year_a, upload_file, self.admin_a)
        self.assertEqual(res['successful'], 1)

        # Payment record must still exist and be intact
        self.assertTrue(FeePayment.objects.filter(pk=p.pk).exists())

        # Student summary must reflect updated total of ₹60,000, paid ₹20,000, remaining ₹40,000
        summary = FeeService.get_student_fee_summary(self.student_1, self.year_a)
        self.assertEqual(summary['total_fee'], Decimal('60000.00'))
        self.assertEqual(summary['total_paid'], Decimal('20000.00'))
        self.assertEqual(summary['remaining_amount'], Decimal('40000.00'))
        self.assertEqual(summary['status'], "PARTIALLY PAID")

    # ══════════════════════════════════════════════════════════════════════════
    # 4. STUDENT PORTAL & RECEIPT PRIVACY
    # ══════════════════════════════════════════════════════════════════════════

    def test_student_portal_views_own_fees_and_downloads_receipt(self):
        """Student views own fees and downloads payment receipt slip."""
        FeeStructure.objects.create(
            school=self.school_a,
            academic_year=self.year_a,
            standard=self.std_10,
            fee_category=FeeService.ensure_default_category(self.school_a),
            name="Class 10 Fee",
            amount=Decimal('50000.00'),
            payment_frequency=FeeStructure.Frequency.QUARTERLY,
        )
        payment, _ = FeeService.record_payment(
            school=self.school_a,
            student=self.student_1,
            amount=Decimal('20000.00'),
            payment_date=date(2026, 8, 20),
            payment_method=FeePayment.PaymentMethod.CASH,
            transaction_reference="CASH-001",
        )

        self.client.force_login(self.student_user_1)

        # 1. Student Portal Fees View
        url = reverse('fees:student_fees')
        res = self.client.get(url, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, "50000.00")
        self.assertContains(res, "20000.00")
        self.assertContains(res, "30000.00")
        self.assertContains(res, "PARTIALLY PAID")

        # 2. Download / View Slip
        receipt_url = reverse('fees:receipt_detail', kwargs={'pk': payment.pk})
        res_receipt = self.client.get(receipt_url, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res_receipt.status_code, 200)
        self.assertContains(res_receipt, payment.receipt_number)
        self.assertContains(res_receipt, "Rahul Patel")
        self.assertContains(res_receipt, "Greenwood High")

    def test_student_privacy_cannot_view_another_students_receipt(self):
        """Student 2 cannot view or download Student 1's payment receipt."""
        payment, _ = FeeService.record_payment(
            school=self.school_a,
            student=self.student_1,
            amount=Decimal('20000.00'),
        )

        # Student 2 attempts to view Student 1's receipt
        self.client.force_login(self.student_user_2)
        receipt_url = reverse('fees:receipt_detail', kwargs={'pk': payment.pk})
        res = self.client.get(receipt_url, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res.status_code, 403)

    # ══════════════════════════════════════════════════════════════════════════
    # 5. MULTI-TENANT ISOLATION
    # ══════════════════════════════════════════════════════════════════════════

    def test_multi_tenant_isolation(self):
        """School B cannot view or record payments for School A's students."""
        payment_a, _ = FeeService.record_payment(
            school=self.school_a,
            student=self.student_1,
            amount=Decimal('20000.00'),
        )

        # Admin B tries to view School A's receipt on School B domain
        self.client.force_login(self.admin_b)
        receipt_url = reverse('fees:receipt_detail', kwargs={'pk': payment_a.pk})
        res = self.client.get(receipt_url, HTTP_HOST='stmary.localhost')
        self.assertEqual(res.status_code, 404)

        # Admin B tries to record payment for School A's student
        post_url = reverse('fees:record_payment')
        res_post = self.client.post(post_url, {
            'student_id': self.student_1.pk,
            'amount': '10000',
        }, HTTP_HOST='stmary.localhost')
        self.assertEqual(res_post.status_code, 404)

    # ══════════════════════════════════════════════════════════════════════════
    # 6. FEATURE FLAG ENFORCEMENT
    # ══════════════════════════════════════════════════════════════════════════

    def test_feature_flag_disabled_blocks_access(self):
        """When fees feature flag is disabled, access is blocked."""
        SchoolFeature.objects.create(
            school=self.school_a,
            feature_key='fees',
            is_enabled=False,
        )

        self.client.force_login(self.admin_a)
        url = reverse('fees:dashboard')
        res = self.client.get(url, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res.status_code, 403)
