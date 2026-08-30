"""
Automated Test Suite for School Timetables.

Covers:
  - Manual timetable creation, editing, clearing
  - Conflict Validation (Faculty double-booking, invalid start/end times)
  - Excel template download
  - Excel bulk upload (multi-class, valid, invalid formats, missing columns, invalid relationships, row-level errors)
  - Class Teacher vs Principal/School Admin permission boundaries
  - Student portal timetable visibility & mutation protection
  - Multi-tenant isolation
  - Feature flag enforcement (enabled vs disabled)
"""
import io
from datetime import date, time
import openpyxl
from django.test import TestCase, Client
from django.urls import reverse
from django.core.files.uploadedfile import SimpleUploadedFile

from apps.accounts.models import User
from apps.academics.models import (
    AcademicYear, Standard, Division, Subject, ClassTimetable, ClassTeacherAllocation
)
from apps.academics.services_timetable import TimetableService, TimetableExcelService
from apps.faculty.models import Faculty
from apps.students.models import Student
from apps.tenants.models import School, SchoolFeature


class TimetableTestSuite(TestCase):
    def setUp(self):
        self.client = Client()

        # 1. School A (Primary Tenant)
        self.school_a = School.objects.create(name="Greenwood High", subdomain="greenwood")

        self.year_a = AcademicYear.objects.create(
            school=self.school_a,
            name="2026-2027",
            start_date=date(2026, 6, 1),
            end_date=date(2027, 5, 31),
            is_current=True,
        )

        self.std_10 = Standard.objects.create(school=self.school_a, name="Std 10", order_index=10)
        self.div_10a = Division.objects.create(school=self.school_a, standard=self.std_10, name="A")
        self.div_10b = Division.objects.create(school=self.school_a, standard=self.std_10, name="B")

        self.std_9 = Standard.objects.create(school=self.school_a, name="Std 9", order_index=9)
        self.div_9a = Division.objects.create(school=self.school_a, standard=self.std_9, name="A")

        self.sub_math = Subject.objects.create(school=self.school_a, name="Mathematics", code="MATH101")
        self.sub_sci = Subject.objects.create(school=self.school_a, name="Science", code="SCI101")
        self.sub_eng = Subject.objects.create(school=self.school_a, name="English", code="ENG101")

        # Admin A
        self.admin_a = User.objects.create_user(
            username="admin_a",
            email="admin@greenwood.edu",
            password="Password@123",
            role=User.Role.SCHOOL_ADMIN,
            school=self.school_a,
        )

        # Faculty 1 (Class Teacher for 10-A)
        self.teacher_user_1 = User.objects.create_user(
            username="rahul_sir",
            email="rahul@greenwood.edu",
            password="Password@123",
            role=User.Role.FACULTY,
            school=self.school_a,
        )
        self.faculty_1 = Faculty.objects.create(
            school=self.school_a,
            user=self.teacher_user_1,
            employee_code="EMP101",
            first_name="Rahul",
            last_name="Sharma",
            email="rahul@greenwood.edu",
        )
        ClassTeacherAllocation.objects.create(
            school=self.school_a,
            academic_year=self.year_a,
            division=self.div_10a,
            faculty=self.faculty_1,
        )

        # Faculty 2 (Class Teacher for 10-B)
        self.teacher_user_2 = User.objects.create_user(
            username="neha_maam",
            email="neha@greenwood.edu",
            password="Password@123",
            role=User.Role.FACULTY,
            school=self.school_a,
        )
        self.faculty_2 = Faculty.objects.create(
            school=self.school_a,
            user=self.teacher_user_2,
            employee_code="EMP102",
            first_name="Neha",
            last_name="Verma",
            email="neha@greenwood.edu",
        )
        ClassTeacherAllocation.objects.create(
            school=self.school_a,
            academic_year=self.year_a,
            division=self.div_10b,
            faculty=self.faculty_2,
        )

        # Faculty 3 (Regular Teacher, no Class Teacher role)
        self.teacher_user_3 = User.objects.create_user(
            username="amit_sir",
            email="amit@greenwood.edu",
            password="Password@123",
            role=User.Role.FACULTY,
            school=self.school_a,
        )
        self.faculty_3 = Faculty.objects.create(
            school=self.school_a,
            user=self.teacher_user_3,
            employee_code="EMP103",
            first_name="Amit",
            last_name="Singh",
            email="amit@greenwood.edu",
        )

        # Student in 10-A
        self.student_user_a = User.objects.create_user(
            username="GR1001",
            email="gr1001@greenwood.edu",
            password="Password@123",
            role=User.Role.STUDENT,
            school=self.school_a,
        )
        self.student_a = Student.objects.create(
            school=self.school_a,
            user=self.student_user_a,
            gr_number="GR1001",
            roll_number=1,
            full_name="Alice Walker",
            academic_year=self.year_a,
            standard=self.std_10,
            division=self.div_10a,
        )

        # 2. School B (Secondary Tenant for Isolation Testing)
        self.school_b = School.objects.create(name="St. Mary Academy", subdomain="stmary")
        self.year_b = AcademicYear.objects.create(
            school=self.school_b,
            name="2026-2027",
            start_date=date(2026, 6, 1),
            end_date=date(2027, 5, 31),
            is_current=True,
        )
        self.std_b = Standard.objects.create(school=self.school_b, name="Grade 10", order_index=10)
        self.div_b = Division.objects.create(school=self.school_b, standard=self.std_b, name="A")
        self.sub_b = Subject.objects.create(school=self.school_b, name="Biology", code="BIO101")
        self.admin_b = User.objects.create_user(
            username="admin_b",
            email="admin@stmary.edu",
            password="Password@123",
            role=User.Role.SCHOOL_ADMIN,
            school=self.school_b,
        )

    # ══════════════════════════════════════════════════════════════════════════
    # 1. MANUAL CREATION & CONFLICT VALIDATION
    # ══════════════════════════════════════════════════════════════════════════

    def test_admin_manual_slot_creation_and_clearing(self):
        """School Admin creates, edits, and clears a timetable slot."""
        self.client.force_login(self.admin_a)

        # 1. Create slot
        url = reverse('academics:timetable_manage')
        post_data = {
            'division_id': self.div_10a.pk,
            'day_of_week': 1,  # Monday
            'period_number': 1,
            'subject_id': self.sub_math.pk,
            'faculty_id': self.faculty_1.pk,
            'start_time': '08:00',
            'end_time': '08:45',
        }
        res = self.client.post(url, post_data, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res.status_code, 302)

        slot = ClassTimetable.objects.get(
            school=self.school_a,
            division=self.div_10a,
            day_of_week=1,
            period_number=1,
        )
        self.assertEqual(slot.subject, self.sub_math)
        self.assertEqual(slot.faculty, self.faculty_1)
        self.assertEqual(slot.start_time, time(8, 0))
        self.assertEqual(slot.end_time, time(8, 45))

        # 2. Clear slot (empty subject_id)
        clear_data = {
            'division_id': self.div_10a.pk,
            'day_of_week': 1,
            'period_number': 1,
            'subject_id': '',
        }
        res_clear = self.client.post(url, clear_data, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res_clear.status_code, 302)
        self.assertFalse(ClassTimetable.objects.filter(
            school=self.school_a,
            division=self.div_10a,
            day_of_week=1,
            period_number=1,
        ).exists())

    def test_invalid_time_sequence_rejected(self):
        """Slot where end_time is before or equal to start_time is rejected."""
        self.client.force_login(self.admin_a)

        url = reverse('academics:timetable_manage')
        post_data = {
            'division_id': self.div_10a.pk,
            'day_of_week': 1,
            'period_number': 1,
            'subject_id': self.sub_math.pk,
            'faculty_id': self.faculty_1.pk,
            'start_time': '09:00',
            'end_time': '08:30',  # Invalid: end before start
        }
        res = self.client.post(url, post_data, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res.status_code, 302)
        # Slot should not be created
        self.assertFalse(ClassTimetable.objects.filter(
            school=self.school_a,
            division=self.div_10a,
            day_of_week=1,
            period_number=1,
        ).exists())

    def test_faculty_double_booking_conflict_rejected(self):
        """Faculty cannot be scheduled in two different classes at the same period."""
        # Book Rahul Sir in 10-A on Monday Period 1
        ClassTimetable.objects.create(
            school=self.school_a,
            academic_year=self.year_a,
            division=self.div_10a,
            day_of_week=1,
            period_number=1,
            subject=self.sub_math,
            faculty=self.faculty_1,
            start_time=time(8, 0),
            end_time=time(8, 45),
        )

        self.client.force_login(self.admin_a)
        url = reverse('academics:timetable_manage')

        # Attempt to book Rahul Sir in 10-B on Monday Period 1
        post_data = {
            'division_id': self.div_10b.pk,
            'day_of_week': 1,
            'period_number': 1,
            'subject_id': self.sub_math.pk,
            'faculty_id': self.faculty_1.pk,
            'start_time': '08:00',
            'end_time': '08:45',
        }
        res = self.client.post(url, post_data, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res.status_code, 302)

        # 10-B slot must NOT have been created
        self.assertFalse(ClassTimetable.objects.filter(
            school=self.school_a,
            division=self.div_10b,
            day_of_week=1,
            period_number=1,
        ).exists())

    # ══════════════════════════════════════════════════════════════════════════
    # 2. CLASS TEACHER ROLE PERMISSIONS
    # ══════════════════════════════════════════════════════════════════════════

    def test_class_teacher_can_manage_assigned_class_only(self):
        """Class Teacher Rahul (10-A) can edit 10-A, but is blocked from editing 10-B."""
        # 1. Rahul edits 10-A -> Allowed
        self.client.force_login(self.teacher_user_1)
        url = reverse('academics:timetable_manage')

        post_data_10a = {
            'division_id': self.div_10a.pk,
            'day_of_week': 2,  # Tuesday
            'period_number': 1,
            'subject_id': self.sub_math.pk,
            'faculty_id': self.faculty_1.pk,
            'start_time': '08:00',
            'end_time': '08:45',
        }
        res_a = self.client.post(url, post_data_10a, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res_a.status_code, 302)
        self.assertTrue(ClassTimetable.objects.filter(
            division=self.div_10a, day_of_week=2, period_number=1
        ).exists())

        # 2. Rahul attempts to edit 10-B -> 403 Forbidden
        post_data_10b = {
            'division_id': self.div_10b.pk,
            'day_of_week': 2,
            'period_number': 1,
            'subject_id': self.sub_sci.pk,
            'faculty_id': self.faculty_2.pk,
        }
        res_b = self.client.post(url, post_data_10b, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res_b.status_code, 403)

    def test_non_class_teacher_faculty_cannot_manage_timetables(self):
        """Faculty without a Class Teacher allocation receives 403 on manage timetable view."""
        self.client.force_login(self.teacher_user_3)  # Amit Sir has no Class Teacher role
        url = reverse('academics:timetable_manage')
        res = self.client.get(url, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res.status_code, 403)

    # ══════════════════════════════════════════════════════════════════════════
    # 3. EXCEL TEMPLATE & BULK UPLOAD
    # ══════════════════════════════════════════════════════════════════════════

    def test_excel_template_download(self):
        """Authorized user can download sample timetable template."""
        self.client.force_login(self.admin_a)
        url = reverse('academics:timetable_template')
        response = self.client.get(url, HTTP_HOST='greenwood.localhost')
        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml.sheet", response['Content-Type'])

        # Verify workbook content
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertIn("Timetable Upload", wb.sheetnames)
        self.assertIn("Reference Masters", wb.sheetnames)

    def _create_test_excel(self, rows, headers=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable Upload"
        if headers is None:
            headers = ["Class", "Day", "Period", "Start Time", "End Time", "Subject", "Faculty"]
        ws.append(headers)
        for r in rows:
            ws.append(r)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_excel_bulk_upload_multi_class_success(self):
        """Admin uploads valid Excel with timetable entries for multiple classes."""
        rows = [
            ["10-A", "Monday", 1, "08:00", "08:45", "Mathematics", "Rahul Sharma"],
            ["10-A", "Monday", 2, "08:45", "09:30", "Science", "Neha Verma"],
            ["10-B", "Monday", 1, "08:00", "08:45", "English", "Amit Singh"],
            ["Std 9-A", "Tuesday", 1, "08:00", "08:45", "Mathematics", "Rahul Sharma"],
        ]
        excel_buf = self._create_test_excel(rows)
        upload_file = SimpleUploadedFile("test_timetable.xlsx", excel_buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        self.client.force_login(self.admin_a)
        url = reverse('academics:timetable_upload')
        res = self.client.post(url, {'excel_file': upload_file}, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res.status_code, 302)

        # Verify all 4 entries are saved in database
        self.assertEqual(ClassTimetable.objects.filter(school=self.school_a).count(), 4)

        # Verify slot in 10-A
        slot_10a_p1 = ClassTimetable.objects.get(division=self.div_10a, day_of_week=1, period_number=1)
        self.assertEqual(slot_10a_p1.subject, self.sub_math)
        self.assertEqual(slot_10a_p1.faculty, self.faculty_1)

        # Verify slot in 10-B
        slot_10b_p1 = ClassTimetable.objects.get(division=self.div_10b, day_of_week=1, period_number=1)
        self.assertEqual(slot_10b_p1.subject, self.sub_eng)
        self.assertEqual(slot_10b_p1.faculty, self.faculty_3)

    def test_excel_upload_invalid_file_type(self):
        """Uploading non-Excel file (.txt / .pdf) returns clear validation error."""
        text_file = SimpleUploadedFile("bad_file.txt", b"Hello World", content_type="text/plain")

        result = TimetableExcelService.import_timetable_excel(
            school=self.school_a,
            academic_year=self.year_a,
            file_obj=text_file,
            user=self.admin_a,
        )
        self.assertEqual(result['total_processed'], 0)
        self.assertEqual(result['successful'], 0)
        self.assertTrue(any("Invalid file type" in e for e in result['errors']))

    def test_excel_upload_missing_columns(self):
        """Excel missing required columns (e.g. missing 'Subject') is rejected."""
        headers = ["Class", "Day", "Period"]  # Missing Subject
        rows = [["10-A", "Monday", 1]]
        excel_buf = self._create_test_excel(rows, headers=headers)
        upload_file = SimpleUploadedFile("missing_col.xlsx", excel_buf.getvalue())

        result = TimetableExcelService.import_timetable_excel(
            school=self.school_a,
            academic_year=self.year_a,
            file_obj=upload_file,
            user=self.admin_a,
        )
        self.assertEqual(result['successful'], 0)
        self.assertTrue(any("Missing required columns" in e for e in result['errors']))

    def test_excel_upload_row_level_validation_errors(self):
        """Excel with non-existent class, subject, faculty produces clear row-level errors."""
        rows = [
            ["NonExistentClass", "Monday", 1, "08:00", "08:45", "Mathematics", "Rahul Sharma"],  # Bad class
            ["10-A", "InvalidDay", 1, "08:00", "08:45", "Mathematics", "Rahul Sharma"],         # Bad day
            ["10-A", "Monday", 99, "08:00", "08:45", "Mathematics", "Rahul Sharma"],           # Bad period
            ["10-A", "Monday", 1, "08:00", "08:45", "NonExistentSubject", "Rahul Sharma"],      # Bad subject
            ["10-A", "Monday", 1, "08:00", "08:45", "Mathematics", "NonExistentTeacher"],      # Bad faculty
        ]
        excel_buf = self._create_test_excel(rows)
        upload_file = SimpleUploadedFile("bad_rows.xlsx", excel_buf.getvalue())

        result = TimetableExcelService.import_timetable_excel(
            school=self.school_a,
            academic_year=self.year_a,
            file_obj=upload_file,
            user=self.admin_a,
        )
        self.assertEqual(result['total_processed'], 5)
        self.assertEqual(result['successful'], 0)
        self.assertEqual(result['failed'], 5)
        self.assertTrue(any("Row 2: Class 'NonExistentClass' not found" in e for e in result['errors']))
        self.assertTrue(any("Row 3: Invalid day" in e for e in result['errors']))
        self.assertTrue(any("Row 4: Invalid period number" in e for e in result['errors']))
        self.assertTrue(any("Row 5: Subject 'NonExistentSubject' not found" in e for e in result['errors']))
        self.assertTrue(any("Row 6: Teacher/Faculty 'NonExistentTeacher' not found" in e for e in result['errors']))

    def test_excel_upload_detects_faculty_double_booking_in_batch(self):
        """Excel containing double-booking of same teacher across classes on same period fails with conflict."""
        rows = [
            ["10-A", "Monday", 1, "08:00", "08:45", "Mathematics", "Rahul Sharma"],
            ["10-B", "Monday", 1, "08:00", "08:45", "Mathematics", "Rahul Sharma"],  # Conflicting with 10-A!
        ]
        excel_buf = self._create_test_excel(rows)
        upload_file = SimpleUploadedFile("conflict.xlsx", excel_buf.getvalue())

        result = TimetableExcelService.import_timetable_excel(
            school=self.school_a,
            academic_year=self.year_a,
            file_obj=upload_file,
            user=self.admin_a,
        )
        self.assertTrue(any("Faculty double-booking in upload" in e for e in result['errors']))

    # ══════════════════════════════════════════════════════════════════════════
    # 4. STUDENT PORTAL & PRIVACY
    # ══════════════════════════════════════════════════════════════════════════

    def test_student_sees_own_class_timetable_and_cannot_modify(self):
        """Student in 10-A views 10-A timetable; cannot access admin manage view or mutate data."""
        # Create slot in 10-A
        ClassTimetable.objects.create(
            school=self.school_a,
            academic_year=self.year_a,
            division=self.div_10a,
            day_of_week=1,
            period_number=1,
            subject=self.sub_math,
            faculty=self.faculty_1,
            start_time=time(8, 0),
            end_time=time(8, 45),
        )

        self.client.force_login(self.student_user_a)

        # 1. View timetable portal -> 200 OK
        url = reverse('academics:student_timetable')
        res = self.client.get(url, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, "Mathematics")
        self.assertContains(res, "Rahul Sharma")

        # 2. Attempt to open manage view -> 403 Forbidden
        manage_url = reverse('academics:timetable_manage')
        res_manage = self.client.get(manage_url, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res_manage.status_code, 403)

    # ══════════════════════════════════════════════════════════════════════════
    # 5. MULTI-TENANT ISOLATION
    # ══════════════════════════════════════════════════════════════════════════

    def test_multi_tenant_isolation_cross_tenant_manipulation_blocked(self):
        """School B Admin cannot view or manipulate School A's timetable data."""
        self.client.force_login(self.admin_b)

        # Attempt to post timetable slot to School A's division
        url = reverse('academics:timetable_manage')
        post_data = {
            'division_id': self.div_10a.pk,  # Belongs to School A
            'day_of_week': 1,
            'period_number': 1,
            'subject_id': self.sub_b.pk,
        }
        res = self.client.post(url, post_data, HTTP_HOST='stmary.localhost')
        self.assertEqual(res.status_code, 404)  # Division not found in School B tenant

    # ══════════════════════════════════════════════════════════════════════════
    # 6. FEATURE FLAG INTEGRATION
    # ══════════════════════════════════════════════════════════════════════════

    def test_feature_flag_disabled_blocks_timetable_access(self):
        """When 'timetable' feature flag is disabled, backend blocks access."""
        # Disable timetable for School A
        SchoolFeature.objects.create(
            school=self.school_a,
            feature_key='timetable',
            is_enabled=False,
        )

        self.client.force_login(self.admin_a)
        url = reverse('academics:timetable_manage')
        res = self.client.get(url, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res.status_code, 403)

        # Student also blocked
        self.client.force_login(self.student_user_a)
        student_url = reverse('academics:student_timetable')
        res_student = self.client.get(student_url, HTTP_HOST='greenwood.localhost')
        self.assertEqual(res_student.status_code, 403)
